"""snapshot_canon.py — canonical text dump of a run's output artifacts.

Used to prove a refactor changed nothing: dump classified.xlsx + summary.json
before and after, and diff the dumps. A byte-compare of the .xlsx itself is
meaningless (the zip container embeds timestamps and openpyxl stamps document
properties), so this emits the *content* deterministically instead:

  - every sheet, in workbook order;
  - every cell that carries information (a value, a solid fill, or a
    hyperlink), row-major, as one line:
        sheet\tcoordinate\tvalue\tfill_rgb\thyperlink
    Fills and hyperlinks are included because the color-coding and the link
    column are load-bearing presentation, not decoration.
  - summary.json re-serialized with sorted keys.

The ONLY field excluded anywhere is summary.json's volatile "generated_at"
timestamp (which also renders into the Run Summary sheet, so the row whose
first cell is "generated_at" is skipped there too). Everything else is
compared verbatim.

Usage:
    python tests/snapshot_canon.py <classified.xlsx> <summary.json> <out.txt>

Standalone on purpose: no gna_pipeline imports, so the same copy runs
unchanged against any checkout.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def _fill_rgb(cell) -> str:
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return "-"
    rgb = getattr(fill.start_color, "rgb", None)
    return str(rgb) if rgb is not None else "-"


def _hyperlink(cell) -> str:
    link = cell.hyperlink
    return str(link.target) if link is not None and link.target else "-"


def dump_workbook(xlsx_path: Path, out) -> None:
    wb = load_workbook(str(xlsx_path))
    for ws in wb.worksheets:
        out.write(f"=== sheet: {ws.title} (dims {ws.dimensions}) ===\n")
        skip_rows_left = 0
        for row in ws.iter_rows():
            # The Run Summary sheet renders summary.json's volatile
            # "generated_at" as a two-row block (header row, then the
            # timestamp value on the next row) — skip both.
            if skip_rows_left:
                skip_rows_left -= 1
                continue
            if row and str(row[0].value) == "generated_at":
                skip_rows_left = 1
                continue
            for cell in row:
                value = cell.value
                fill = _fill_rgb(cell)
                link = _hyperlink(cell)
                if value is None and fill == "-" and link == "-":
                    continue
                out.write(f"{ws.title}\t{cell.coordinate}\t{value!r}\t{fill}\t{link}\n")


def dump_summary(summary_path: Path, out) -> None:
    data = json.loads(Path(summary_path).read_text(encoding="utf-8"))
    data.pop("generated_at", None)
    out.write("=== summary.json (generated_at excluded) ===\n")
    out.write(json.dumps(data, indent=2, sort_keys=True, default=str))
    out.write("\n")


def main(argv: list[str]) -> int:
    if len(argv) != 4:
        print(__doc__, file=sys.stderr)
        return 2
    xlsx_path, summary_path, out_path = Path(argv[1]), Path(argv[2]), Path(argv[3])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8", newline="\n") as out:
        dump_workbook(xlsx_path, out)
        dump_summary(summary_path, out)
    print(f"wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
