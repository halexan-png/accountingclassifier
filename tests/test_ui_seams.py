"""tests/test_ui_seams.py — Wave 1A backend seams for the G&A Classifier UI.

Covers the additive, CLI-byte-identical seams added for `gna_server` (v2 UI
handoff §6): console.py's event sink / structured data channel / pluggable
confirm / Progress.snapshot(); scheduling.py's cooperative OperatorCancelled
cancel; pipeline.run_pipeline's `user_deal_context_override` kwarg; and
ingest's workbook-validation callable. (Wave 1 removed resume/reuse entirely,
including excel_out's former resume-by-reupload translator -- there is no
longer anything in that category to cover here.) Also covers Stream B's
company-norms context plumbing (prompts.load_company_norms + its optional
system block).

None of this is exercised by the CLI today (every new parameter defaults to
"off"/None), so these tests prove the NEW surface directly rather than via
cli.cmd_run.
"""

from __future__ import annotations

import itertools
import threading
import time

import pytest
from openpyxl import Workbook

from gna_pipeline import (
    classify,
    config,
    console,
    deal_profile,
    ingest,
    pipeline,
    prep,
    scheduling,
)
from gna_pipeline.contract import make_decision_record, zero_usage


# ---------------------------------------------------------------------------
# console.py — event sink
# ---------------------------------------------------------------------------

@pytest.fixture
def sink(monkeypatch):
    """Install a recording event sink for the test, and guarantee console.py's
    module-level state (sink, confirm handler, status line) is reset after —
    these are process-global, so a leaked sink/handler would bleed into
    unrelated tests."""
    events: list[tuple[str, dict]] = []
    console.set_event_sink(lambda kind, payload: events.append((kind, payload)))
    try:
        yield events
    finally:
        console.set_event_sink(None)
        console.set_confirm_handler(None)
        console.clear_status()


def test_section_banner_info_warn_kv_mirror_to_sink(sink, capsys):
    console.section("Ingest")
    console.banner(["line one", "line two"])
    console.info("hello")
    console.warn("uh oh")
    console.kv({"a": "1", "b": "2"})

    out = capsys.readouterr().out
    assert "Ingest" in out and "hello" in out and "uh oh" in out

    kinds = [k for k, _ in sink]
    assert kinds == ["section", "banner", "info", "warn", "kv"]
    assert sink[0][1] == {"title": "Ingest"}
    assert sink[1][1] == {"lines": ["line one", "line two"]}
    assert sink[2][1] == {"msg": "hello"}
    assert sink[3][1] == {"msg": "uh oh"}
    assert sink[4][1] == {"pairs": [["a", "1"], ["b", "2"]]}


def test_row_mirrors_row_shape_to_sink(sink):
    record = make_decision_record(
        packet={"row_idx": 42, "acctnum": "GA10000000", "amount": 123.45, "currency": "USD"},
        row_hash="deadbeef",
        phase="classify",
        classification="human_review",
        basis="none",
        reasoning="needs a human",
        evidence="[row] ...",
        missing_info="what deal is this",
        flags=["amount_blank"],
    )
    console.row(record, "classify")

    assert len(sink) == 1
    kind, payload = sink[0]
    assert kind == "row"
    assert payload == {
        "row_idx": 42,
        "acctnum": "GA10000000",
        "amount": 123.45,
        "currency": "USD",
        "classification": "human_review",
        "basis": "none",
        "recognized_deal": "none",
        "flags": ["amount_blank"],
        "phase": "classify",
        "missing_info": "what deal is this",
        "error": None,
    }


def test_data_channel_prints_nothing_sink_only(sink, capsys):
    console.data("phase0_stats", {"reclass_fired": 3})
    out = capsys.readouterr().out
    assert out == ""
    assert sink == [("data", {"kind": "phase0_stats", "payload": {"reclass_fired": 3}})]


def test_status_and_clear_status_forward_snapshot(sink):
    console.status("42/100 rows", snapshot={"done_rows": 42, "total_rows": 100})
    console.clear_status()

    assert sink[0] == ("status", {"text": "42/100 rows", "snapshot": {"done_rows": 42, "total_rows": 100}})
    assert sink[1] == ("status", {"text": "", "snapshot": None})


def test_sink_exception_never_breaks_a_run(capsys):
    def bad_sink(kind, payload):
        raise RuntimeError("boom")

    console.set_event_sink(bad_sink)
    try:
        console.info("still works")  # must not raise
    finally:
        console.set_event_sink(None)
    assert "still works" in capsys.readouterr().out


# ---------------------------------------------------------------------------
# pipeline.py print helpers actually emit their data() events (§6.0 item 2 —
# a code-review pass on this wave found console.data() defined but never
# called anywhere in production code; these tests lock the fix down so a
# future refactor can't silently drop the wiring again).
# ---------------------------------------------------------------------------

def test_print_phase0_stats_emits_data_event(sink):
    stats0 = prep._empty_stats()
    pipeline.print_phase0_stats(stats0)
    # `work_items` (internal; can carry raw PDF bytes) is stripped from the
    # event -- only the scalar counts are streamed to the UI.
    expected = {k: v for k, v in stats0.items() if k != "work_items"}
    assert ("data", {"kind": "phase0_stats", "payload": expected}) in sink


def test_print_phase1_forecast_emits_data_event(sink):
    forecast = {
        "rows": 1, "est_batches": 1, "input_tokens_low": 100, "input_tokens_high": 100,
        "cost_low_usd": 0.01, "cost_high_usd": 0.01, "max_workers": 1, "wall_clock_est_min": 0.1,
    }
    pipeline.print_phase1_forecast(forecast)
    assert ("data", {"kind": "sweep_forecast", "payload": forecast}) in sink


def test_print_phase2_forecast_emits_data_event(sink):
    forecast = {
        "rows": 1, "rows_with_invoice": 0, "est_batches": 1, "input_tokens_low": 100,
        "input_tokens_high": 100, "output_tokens_est": 10, "cost_low_usd": 0.01,
        "cost_high_usd": 0.01, "max_workers": 1, "wall_clock_est_min": 0.1,
    }
    pipeline._print_phase2_forecast(forecast, None)
    assert ("data", {"kind": "classify_forecast", "payload": forecast}) in sink


def test_print_context_report_emits_data_event(sink):
    report = {
        "est_tokens": 0, "cap": 4000, "cap_quarters": 1,
        "entries_full": 0, "collapsed": [], "dropped": [],
    }
    pipeline.print_context_report(report)
    assert ("data", {"kind": "context_report", "payload": report}) in sink


def test_print_closing_tally_emits_data_event(sink):
    summary = {"tally": {"by_classification": {}}, "usage": {"cost_actual_usd": 0.0}}
    pipeline.print_closing_tally(summary)
    assert ("data", {"kind": "closing_tally", "payload": summary}) in sink


def test_stage9_and_stage10_emit_phase_stepper_events(sink, monkeypatch):
    pipeline._stage9_phase1_sweep(
        row_mode=True, ma_scope_packets=[], selected=[],
        excluded=[], quarters=[], model=config.DEFAULT_MODEL, human_deals_md=None,
        client=object(), rate_limits=None, sweep_fc={}, yes=True,
    )
    assert ("data", {"kind": "phase", "payload": {"phase": "build_profile"}}) in sink

    monkeypatch.setattr(pipeline, "write_context_txt", lambda *_a, **_k: None)
    monkeypatch.setattr(classify, "run_classification", lambda *a, **k: [])
    pipeline._stage10_phase2_classify(
        interrupted=False, declined=False, row_mode=False, profile=pipeline.empty_deal_profile(),
        human_deals_md=None, classify_items=[],
        model=config.DEFAULT_MODEL, rate_limits=None,
        forecast={"cost_high_usd": 1.0}, client=object(), stats0=prep._empty_stats(),
    )
    assert ("data", {"kind": "phase", "payload": {"phase": "classify"}}) in sink


def test_stage11_emits_outputs_and_done_phase_events(sink, monkeypatch, tmp_path):
    monkeypatch.setattr(config, "SUMMARY_JSON", tmp_path / "summary.json")
    monkeypatch.setattr(config, "CLASSIFIED_XLSX", tmp_path / "classified.xlsx")

    source = tmp_path / "source.xlsx"
    _write_test_source(source, [_row(ref="R1")])
    packets, _stats = ingest.read_packets(str(source), sheet=config.SHEET_NAME)

    rc = pipeline._stage11_output(
        workbook=source, packets=packets, phase0_records=[],
        sweep_records=[], classify_results=[], forecast={}, profile=pipeline.empty_deal_profile(),
        scope_stats={}, run_start=time.monotonic(), interrupted=False, declined=False,
    )
    assert rc == 0

    data_events = [payload for kind, payload in sink if kind == "data"]
    assert any(d["kind"] == "outputs" and d["payload"]["excel_ok"] is True for d in data_events)
    assert {"kind": "phase", "payload": {"phase": "done"}} in data_events


# ---------------------------------------------------------------------------
# console.py — pluggable confirm
# ---------------------------------------------------------------------------

def test_confirm_uses_handler_when_set():
    calls = []

    def handler(prompt):
        calls.append(prompt)
        return True

    console.set_confirm_handler(handler)
    try:
        assert console.confirm("Proceed? [y/N] ") is True
    finally:
        console.set_confirm_handler(None)
    assert calls == ["Proceed? [y/N] "]


def test_confirm_falls_back_to_input_when_no_handler(monkeypatch):
    monkeypatch.setattr("builtins.input", lambda prompt: "y")
    assert console.confirm("Proceed? [y/N] ") is True


def test_confirm_declines_on_eof_never_auto_yes(monkeypatch):
    """Stdin not interactive (no tty, redirected from /dev/null, etc.) must
    decline the money gate, not crash on the EOFError traceback and never
    silently assume "yes"."""
    def raise_eof(prompt):
        raise EOFError()

    monkeypatch.setattr("builtins.input", raise_eof)
    assert console.confirm("Proceed? [y/N] ") is False


# ---------------------------------------------------------------------------
# console.py — Progress.snapshot()
# ---------------------------------------------------------------------------

def test_progress_snapshot_shape_and_render_agree():
    progress = console.Progress(total_rows=10, total_batches=2, model=config.DEFAULT_MODEL, unit="rows")
    progress.batch_started()
    progress.batch_done(records=[
        {"usage": {"input_tokens": 1000, "output_tokens": 100,
                    "cache_read_input_tokens": 0, "cache_creation_input_tokens": 0}}
        for _ in range(5)
    ])

    snap = progress.snapshot()
    assert snap.keys() == {
        "done_rows", "total_rows", "done_batches", "total_batches",
        "in_flight", "cost_usd", "eta_s", "unit",
    }
    assert snap["done_rows"] == 5
    assert snap["total_rows"] == 10
    assert snap["done_batches"] == 1
    assert snap["total_batches"] == 2
    assert snap["in_flight"] == 0
    assert snap["unit"] == "rows"
    assert snap["eta_s"] is not None and snap["eta_s"] >= 0

    rendered = progress.render()
    assert f"{snap['done_rows']}/{snap['total_rows']}" in rendered
    assert f"${snap['cost_usd']:.2f}" in rendered


def test_progress_snapshot_eta_none_before_any_batch_completes():
    progress = console.Progress(total_rows=10, total_batches=2, model=config.DEFAULT_MODEL)
    assert progress.snapshot()["eta_s"] is None
    assert "ETA --" in progress.render()


# ---------------------------------------------------------------------------
# scheduling.py — OperatorCancelled / cooperative cancel
# ---------------------------------------------------------------------------

def test_operator_cancelled_subclasses_keyboard_interrupt():
    assert issubclass(scheduling.OperatorCancelled, KeyboardInterrupt)
    assert not issubclass(scheduling.OperatorCancelled, Exception)


@pytest.fixture
def cancel_event():
    ev = threading.Event()
    scheduling.set_cancel_event(ev)
    try:
        yield ev
    finally:
        scheduling.set_cancel_event(None)


def test_run_batches_raises_operator_cancelled_and_stops_queued_batches(cancel_event, monkeypatch):
    warnings = []
    monkeypatch.setattr(console, "warn", lambda msg: warnings.append(msg))

    processed = []

    def process_one(batch):
        processed.append(batch)
        return batch

    def on_complete(result):
        pass

    cancel_event.set()  # already cancelled before the run even starts
    batches = [["warmup"], ["b1"], ["b2"], ["b3"]]

    with pytest.raises(scheduling.OperatorCancelled):
        scheduling.run_batches(
            batches, process_one, max_workers=2, on_complete=on_complete,
            interrupt_label="test-phase",
        )

    # The warmup batch (batches[0]) always runs synchronously regardless of
    # cancel state (existing, unchanged behavior); no *queued* batch beyond
    # it should have been submitted once cancel was already set.
    assert processed == [["warmup"]]
    assert any("interrupted" in w for w in warnings)


def test_rate_limiter_acquire_raises_operator_cancelled_before_admission(cancel_event):
    limiter = scheduling.RateLimiter(rpm_cap=100, itpm_cap=None, otpm_cap=None)
    cancel_event.set()
    with pytest.raises(scheduling.OperatorCancelled):
        limiter.acquire(input_tokens=1, output_tokens=1)


def test_operator_cancelled_caught_like_keyboard_interrupt_in_phase2(monkeypatch):
    """Mirrors the existing except (KeyboardInterrupt, SpendCapExceeded) site
    in pipeline._stage10_phase2_classify: an OperatorCancelled raised out of
    classify.run_classification must be caught there (it subclasses
    KeyboardInterrupt) and reported as `interrupted`, not propagate."""
    monkeypatch.setattr(pipeline, "write_context_txt", lambda *_a, **_k: None)

    def fake_run_classification(*args, **kwargs):
        raise scheduling.OperatorCancelled("cancelled by UI Stop")

    monkeypatch.setattr(classify, "run_classification", fake_run_classification)

    stats0 = prep._empty_stats()
    profile = pipeline.empty_deal_profile()
    packet = {
        "row_idx": 5, "period": "202601", "acctnum": "GA10000000", "ref": "", "entityid": "",
        "department": None, "category": None, "amount": 100.0, "amount_was_blank": False,
        "currency": "USD", "descrptn": "vendor fee", "adddesc": None, "source": "AP",
        "entrdate": None, "userid": "jdoe", "invoice_url": None,
    }
    work_item = {
        "packet": packet, "row_hash": "h", "had_invoice": "no", "invoice_accessed": "no",
        "flags": [], "invoice": None, "est_input_tokens": 100,
    }

    classify_results, interrupted = pipeline._stage10_phase2_classify(
        interrupted=False, declined=False, row_mode=False, profile=profile,
        human_deals_md=None, classify_items=[work_item],
        model=config.DEFAULT_MODEL, rate_limits=None,
        forecast={"cost_high_usd": 1.0}, client=object(), stats0=stats0,
    )

    assert interrupted is True
    assert classify_results == []


# ---------------------------------------------------------------------------
# pipeline.py — user_deal_context_override (§6.0 item 8)
# ---------------------------------------------------------------------------

def _stub_run_pipeline_through_forecast(monkeypatch, recorder):
    """Stub every stage function so run_pipeline reaches the
    human_deals_md-resolution line and the Phase-1/Phase-2 stage calls
    without touching disk, the network, or a real API client."""
    monkeypatch.setattr(pipeline, "_stage1_ingest_and_scope", lambda *a, **k: ([], {}))
    monkeypatch.setattr(pipeline, "_stage2_select_quarters", lambda *a, **k: ([], []))
    monkeypatch.setattr(pipeline, "_stage4_load_lookup_index", lambda: {})
    monkeypatch.setattr(pipeline, "_stage5_phase0", lambda *a, **k: ([], {}, [], None))
    monkeypatch.setattr(pipeline, "_stage6_partition", lambda *a, **k: ([], [], [], []))
    monkeypatch.setattr(pipeline, "_stage7_forecast", lambda *a, **k: ({}, {}))
    monkeypatch.setattr(pipeline, "need_api_key", lambda: False)
    monkeypatch.setattr(pipeline, "build_client", lambda: object())
    monkeypatch.setattr(pipeline, "refresh_rate_limits", lambda *a, **k: None)

    def fake_stage9(**kwargs):
        recorder["stage9_human_deals_md"] = kwargs["human_deals_md"]
        return ({}, [], False, False)

    def fake_stage10(**kwargs):
        recorder["stage10_human_deals_md"] = kwargs["human_deals_md"]
        return ([], False)

    monkeypatch.setattr(pipeline, "_stage9_phase1_sweep", fake_stage9)
    monkeypatch.setattr(pipeline, "_stage10_phase2_classify", fake_stage10)
    monkeypatch.setattr(pipeline, "_stage11_output", lambda **k: 0)


def _run(**overrides):
    base = dict(
        workbook="dummy.xlsx", model=None, n=None, rows=None, quarters=None, months=None,
        min_usd=None, dry_run=False, yes=True, no_fetch=True,
    )
    base.update(overrides)
    return pipeline.run_pipeline(**base)


def test_user_deal_context_override_wins_over_disk(monkeypatch):
    from gna_pipeline import deal_profile, prompts

    monkeypatch.setattr(deal_profile, "load_human_deals_md", lambda: "FROM_DISK")

    recorder: dict = {}
    _stub_run_pipeline_through_forecast(monkeypatch, recorder)

    rc = _run(user_deal_context_override="FROM_UI_OVERRIDE")
    assert rc == 0
    assert recorder["stage9_human_deals_md"] == "FROM_UI_OVERRIDE"
    assert recorder["stage10_human_deals_md"] == "FROM_UI_OVERRIDE"


def test_user_deal_context_omitted_is_byte_identical_to_today(monkeypatch):
    from gna_pipeline import deal_profile, prompts

    monkeypatch.setattr(deal_profile, "load_human_deals_md", lambda: "FROM_DISK")

    recorder: dict = {}
    _stub_run_pipeline_through_forecast(monkeypatch, recorder)

    rc = _run()  # user_deal_context_override omitted -> default None
    assert rc == 0
    assert recorder["stage9_human_deals_md"] == "FROM_DISK"
    assert recorder["stage10_human_deals_md"] == "FROM_DISK"


def test_human_deals_md_lands_in_classify_system_blocks():
    """The tests above stop at 'the kwarg reaches _stage10_phase2_classify' --
    this closes the next link: does that stage's own prompt builder
    (prompts.build_system_prompt) actually put the text into a system block?
    Mirrors pipeline.py:834's real call shape."""
    from gna_pipeline import prompts

    blocks = prompts.build_system_prompt(None, "MARKER_HUMAN_CONTEXT_TEXT")
    texts = [b["text"] for b in blocks]
    assert any("MARKER_HUMAN_CONTEXT_TEXT" in t for t in texts)
    # and it's tagged with the Phase-2 top-authority header (mandate framing),
    # not silently merged in
    assert any("OPERATOR CONTEXT" in t and "MARKER_HUMAN_CONTEXT_TEXT" in t for t in texts)
    # and the hardcoded order-of-authority block is always present
    assert any("=== ORDER OF AUTHORITY ===" in t for t in texts)


def test_human_deals_md_lands_in_sweep_system_blocks():
    """Same link, Phase 1 (deal_profile.run_sweep's prompt builder)."""
    from gna_pipeline import prompts

    blocks = prompts.build_sweep_system_prompt("MARKER_HUMAN_CONTEXT_TEXT")
    texts = [b["text"] for b in blocks]
    assert any("MARKER_HUMAN_CONTEXT_TEXT" in t for t in texts)


class _FakeMessages:
    def __init__(self, calls: list[dict]) -> None:
        self._calls = calls

    def create(self, **kwargs):
        self._calls.append(kwargs)
        return "fake-response"


class _FakeAnthropicClient:
    def __init__(self) -> None:
        self.calls: list[dict] = []
        self.messages = _FakeMessages(self.calls)


def test_human_deals_md_reaches_the_raw_api_request():
    """The last link: given the system blocks prompts.py built (containing
    the human-authored text), does classify._make_request actually hand them
    to `client.messages.create(system=...)` -- the literal outbound Anthropic
    API call? This is the same private function pipeline.py's Phase 2 uses
    for every batch."""
    from gna_pipeline import prompts

    system_blocks = prompts.build_system_prompt(None, "MARKER_HUMAN_CONTEXT_TEXT")
    packet = {
        "row_idx": 5, "period": "202601", "acctnum": "GA10000000", "ref": "", "entityid": "",
        "department": None, "category": None, "amount": 100.0, "amount_was_blank": False,
        "currency": "USD", "descrptn": "vendor fee", "adddesc": None, "source": "AP",
        "entrdate": None, "userid": "jdoe", "invoice_url": None,
    }
    work_item = {
        "packet": packet, "row_hash": "h", "had_invoice": "no", "invoice_accessed": "no",
        "flags": [], "invoice": None, "est_input_tokens": 100,
    }
    client = _FakeAnthropicClient()

    classify._make_request(client, [work_item], system_blocks, config.DEFAULT_MODEL)

    assert len(client.calls) == 1
    sent_system = client.calls[0]["system"]
    assert any("MARKER_HUMAN_CONTEXT_TEXT" in b.get("text", "") for b in sent_system)


# ---------------------------------------------------------------------------
# Two-tier model policy: config.model_for_batch's routing, classify._make_
# request's disabled-thinking kwarg, per-batch model_version stamping
# through classify._process_batch (scripted client, no network), and
# classify.forecast's per-tier cost split.
# ---------------------------------------------------------------------------

def _tiering_packet(row_idx: int) -> dict:
    return {
        "row_idx": row_idx, "period": "202601", "acctnum": "GA10000000", "ref": "", "entityid": "",
        "department": None, "category": None, "amount": 100.0, "amount_was_blank": False,
        "currency": "USD", "descrptn": "vendor fee", "adddesc": None, "source": "AP",
        "entrdate": None, "userid": "jdoe", "invoice_url": None,
    }


def _tiering_work_item(row_idx: int, *, invoice=None, est_input_tokens=100) -> dict:
    return {
        "packet": _tiering_packet(row_idx), "row_hash": f"h{row_idx}", "had_invoice": "no",
        "invoice_accessed": "no", "flags": [], "invoice": invoice,
        "est_input_tokens": est_input_tokens,
    }


def test_model_for_batch_invoice_batch_runs_on_invoice_model():
    batch = [
        _tiering_work_item(1, invoice=None),
        _tiering_work_item(2, invoice={"kind": "pdf"}),
    ]
    assert config.model_for_batch(batch) == config.INVOICE_MODEL


def test_model_for_batch_no_invoice_runs_on_floor():
    batch = [_tiering_work_item(1, invoice=None), _tiering_work_item(2, invoice={"kind": "none"})]
    assert config.model_for_batch(batch) == config.DEFAULT_MODEL


def test_model_for_batch_floor_override_respected():
    floor_batch = [_tiering_work_item(1, invoice=None)]
    assert config.model_for_batch(floor_batch, floor_model="claude-custom-floor") == "claude-custom-floor"
    # the override only ever changes the FLOOR -- an invoice batch still upgrades.
    invoice_batch = [_tiering_work_item(1, invoice={"kind": "text"})]
    assert (
        config.model_for_batch(invoice_batch, floor_model="claude-custom-floor")
        == config.INVOICE_MODEL
    )


def test_config_scope_and_worker_ceiling_defaults():
    assert config.SCOPE_MIN_USD_DEFAULT == 999.0
    assert config.MAX_WORKERS_CEILING == 30


def test_make_request_disables_thinking():
    """Every classify request disables adaptive thinking so the forced-tool
    extraction stays deterministic (classify._make_request's `thinking`
    kwarg) -- a no-op for models without it, load-bearing for Sonnet 5."""
    client = _FakeAnthropicClient()
    work_item = _tiering_work_item(1)

    classify._make_request(client, [work_item], [], config.DEFAULT_MODEL)

    assert client.calls[0]["thinking"] == {"type": "disabled"}


class _FakeToolBlock:
    def __init__(self, rows):
        self.type = "tool_use"
        self.name = "classify_rows"
        self.input = {"rows": rows}


class _FakeUsage:
    def __init__(self):
        self.input_tokens = 10
        self.output_tokens = 5
        self.cache_read_input_tokens = 0
        self.cache_creation_input_tokens = 0


class _FakeToolResponse:
    def __init__(self, rows):
        self.content = [_FakeToolBlock(rows)]
        self.usage = _FakeUsage()
        self.stop_reason = "tool_use"


class _ScriptedMessages:
    def __init__(self, rows):
        self._rows = rows
        self.calls: list[dict] = []

    def create(self, **kwargs):
        self.calls.append(kwargs)
        return _FakeToolResponse(self._rows)


class _ScriptedClient:
    """Fake Anthropic client for `_process_batch`: scripted to return a
    valid classify_rows tool response for exactly the row_idx(s) given, so
    the anti-conflation check passes and the batch's own model routing (not
    the network) is what's under test."""

    def __init__(self, row_idxs):
        rows = [
            {
                "row_idx": ri, "classification": "recurring", "basis": "row_text_routine",
                "reasoning": "r", "evidence": "e", "recognized_deal": "none",
                "invoice_read": "none_attached",
                "missing_info": None,
            }
            for ri in row_idxs
        ]
        self.messages = _ScriptedMessages(rows)


def test_process_batch_stamps_per_batch_model_version():
    """End to end through `_process_batch` (no network): a batch holding a
    readable invoice stamps every record's model_version with
    config.INVOICE_MODEL; a batch with none stamps the floor -- proves
    config.model_for_batch's routing reaches the actual DecisionRecord, not
    just the outbound `model=` kwarg."""
    limiter = scheduling.build_limiter(None)
    emitted: list = []

    invoice_batch = [_tiering_work_item(1, invoice={"kind": "text", "read_path": "text"})]
    invoice_client = _ScriptedClient([1])
    invoice_records = classify._process_batch(
        invoice_client, invoice_batch, [], config.DEFAULT_MODEL, limiter,
        emitted.append, itertools.count(1),
    )
    assert invoice_records[0]["model_version"] == config.INVOICE_MODEL
    assert invoice_client.messages.calls[0]["model"] == config.INVOICE_MODEL

    floor_batch = [_tiering_work_item(2, invoice=None)]
    floor_client = _ScriptedClient([2])
    floor_records = classify._process_batch(
        floor_client, floor_batch, [], config.DEFAULT_MODEL, limiter,
        emitted.append, itertools.count(1),
    )
    assert floor_records[0]["model_version"] == config.DEFAULT_MODEL
    assert floor_client.messages.calls[0]["model"] == config.DEFAULT_MODEL


# ---------------------------------------------------------------------------
# Vision-failure -> text/row-only fallback (both phases). A failed native-PDF
# vision request does NOT fail the row: the invoice is downgraded and the row
# is re-run ONCE on the floor model from its own information.
# ---------------------------------------------------------------------------

class _FailFirstThenToolMessages:
    """messages.create raises on the FIRST call (the vision attempt) and returns
    `response_factory()` on every call after (the text/row-only fallback)."""

    def __init__(self, response_factory):
        self._response_factory = response_factory
        self.calls: list[dict] = []

    def create(self, **kwargs):
        self.calls.append(kwargs)
        if len(self.calls) == 1:
            raise RuntimeError("vision boom")
        return self._response_factory()


class _FailFirstClient:
    def __init__(self, response_factory):
        self.messages = _FailFirstThenToolMessages(response_factory)

    def with_options(self, **kwargs):  # vision path caps retries via this
        return self


def _vision_invoice(url="http://x/inv.pdf"):
    return {
        "kind": "pdf", "read_path": "vision", "pdf_bytes": b"%PDF-1.4 fake",
        "source": "url", "path_or_url": url, "sha256": "abc", "size_bytes": 10,
        "text": "", "pages_read": "1-1", "page_count": 1, "error": None,
    }


def test_classify_vision_failure_falls_back_to_text_row_only():
    limiter = scheduling.build_limiter(None)
    emitted: list = []
    item = _tiering_work_item(1, invoice=_vision_invoice())
    item["had_invoice"] = "yes"
    item["invoice_accessed"] = "yes"
    client = _FailFirstClient(lambda: _FakeToolResponse([{
        "row_idx": 1, "classification": "recurring", "basis": "row_text_routine",
        "reasoning": "r", "evidence": "e", "recognized_deal": "none",
        "invoice_read": "unavailable",
        "missing_info": None,
    }]))

    records = classify._process_batch(
        client, [item], [], config.DEFAULT_MODEL, limiter, emitted.append, itertools.count(1),
    )

    # Exactly two create() calls: the vision attempt (raised, no retry) and the
    # text/row-only fallback (succeeded).
    assert len(client.messages.calls) == 2
    assert client.messages.calls[0]["model"] == config.INVOICE_MODEL  # vision -> stronger model
    assert client.messages.calls[1]["model"] == config.DEFAULT_MODEL   # fallback -> floor
    # The row is classified, not error-recorded.
    assert records[0]["error"] is None
    assert records[0]["decision"]["classification"] == "recurring"
    # The record still names WHICH invoice failed, and why.
    assert records[0]["invoice"]["kind"] == "error"
    assert "vision boom" in (records[0]["invoice"]["error"] or "")


class _FakeSweepToolBlock:
    def __init__(self, entries):
        self.type = "tool_use"
        self.name = "build_deal_profile"
        self.input = {"entries": entries}


class _FakeSweepResponse:
    def __init__(self, entries):
        self.content = [_FakeSweepToolBlock(entries)]
        self.usage = _FakeUsage()
        self.stop_reason = "tool_use"


def _ma_vision_work_item(row_idx: int) -> dict:
    packet = _tiering_packet(row_idx)
    packet["acctnum"] = config.MA_ACCTNUM
    return {
        "packet": packet, "row_hash": f"h{row_idx}", "had_invoice": "yes",
        "invoice_accessed": "yes", "flags": [], "invoice": _vision_invoice(),
        "est_input_tokens": 100,
    }


def test_sweep_vision_failure_falls_back_to_text_row_only():
    limiter = scheduling.build_limiter(None)
    lock = threading.Lock()
    emitted: list = []
    entries_acc: dict = {}
    stats = {
        "rows_selected": 1, "rows_ok": 0, "rows_failed": 0, "invoices_read": 1,
        "batches_total": 1, "batches_failed": 0, "entries": 0,
    }
    usage_total = zero_usage()
    cost_total: dict = {}
    item = _ma_vision_work_item(1)
    client = _FailFirstClient(lambda: _FakeSweepResponse([
        {"name": "Project Apollo", "evidence": ["[invoice] acquisition of Apollo"]}
    ]))

    batch_failed = deal_profile._process_sweep_batch(
        client, [item], [], config.DEFAULT_MODEL, limiter, lock, emitted.append,
        entries_acc, stats, usage_total, cost_total,
    )

    # Two create() calls: the vision attempt (raised) then the text/row-only
    # fallback (succeeded), so the row swept OK, NOT deal_sweep_failed.
    assert len(client.messages.calls) == 2
    assert client.messages.calls[0]["model"] == config.INVOICE_MODEL
    assert client.messages.calls[1]["model"] == config.DEFAULT_MODEL
    assert batch_failed is False
    assert stats["rows_ok"] == 1
    assert stats["rows_failed"] == 0
    assert "deal_sweep_failed" not in emitted[-1]["flags"]
    assert emitted[-1]["decision"]["classification"] == "non_recurring"


def test_forecast_splits_cost_by_tier():
    """classify.forecast prices invoice-bearing rows at config.INVOICE_MODEL
    and everything else at the floor, and exposes both tiers in cost_by_tier
    alongside the blended cost_low_usd/cost_high_usd total."""
    items = [
        _tiering_work_item(
            1, invoice={"kind": "text", "read_path": "text", "text": "x" * 4000,
                        "est_input_tokens": 1000},
            est_input_tokens=1500,
        ),
        _tiering_work_item(
            2, invoice={"kind": "text", "read_path": "text", "text": "x" * 4000,
                        "est_input_tokens": 1000},
            est_input_tokens=1500,
        ),
        _tiering_work_item(3, invoice=None, est_input_tokens=200),
        _tiering_work_item(4, invoice=None, est_input_tokens=200),
    ]

    fc = classify.forecast(items, config.DEFAULT_MODEL, None)

    assert fc["rows_with_invoice"] == 2
    assert fc["cost_by_tier"]["invoice"]["model"] == config.INVOICE_MODEL
    assert fc["cost_by_tier"]["invoice"]["rows"] == 2
    assert fc["cost_by_tier"]["floor"]["model"] == config.DEFAULT_MODEL
    assert fc["cost_by_tier"]["floor"]["rows"] == 2
    assert fc["cost_by_tier"]["invoice"]["cost_high_usd"] > 0
    assert fc["cost_by_tier"]["floor"]["cost_high_usd"] > 0
    assert fc["cost_low_usd"] <= fc["cost_high_usd"]
    assert fc["cost_low_usd"] == pytest.approx(
        fc["cost_by_tier"]["invoice"]["cost_low_usd"] + fc["cost_by_tier"]["floor"]["cost_low_usd"]
    )


# ---------------------------------------------------------------------------
# ingest.validate_workbook (§6.1)
# ---------------------------------------------------------------------------

# A clean 20-column header row: every field's alias matches at its own
# fallback position, so ingest.resolve_columns reports zero warnings.
_HEADERS = [
    "Period", "Ref", "Source", "EntityID", "AcctNum", "Department", "Unused7",
    "Descrpn", "Unused9", "EntrDate", "OCurrCode", "AddlDesc", "Unused13",
    "UserID", "Category", "Unused16", "Unused17", "Image URL - Hyperlink",
    "Unused19", "USD Amount",
]


def _row(*, ref, acctnum="GA10000000", descrptn="Routine vendor invoice", amount=100.0, userid="jdoe"):
    return [
        "202601", ref, "AP", "ENT1", acctnum, "DEPT1", None,
        descrptn, None, "2026-01-15", "USD", None, None,
        userid, "Vendor", None, None, None,
        None, amount,
    ]


def _write_test_source(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = config.SHEET_NAME[:31]
    for col, header in enumerate(_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    for r, row_values in enumerate(rows, start=2):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=r, column=col, value=value)
    wb.save(str(path))


def test_validate_workbook_clean_file_all_checks_pass(tmp_path):
    source = tmp_path / "source.xlsx"
    _write_test_source(source, [_row(ref="R1"), _row(ref="R2")])

    result = ingest.validate_workbook(str(source), sheet=config.SHEET_NAME)

    assert result["row_count"] == 2
    assert result["has_existing_classifications"] is False
    assert {c["label"]: c["ok"] for c in result["checks"]} == {
        "File format": True, "Worksheet structure": True, "Expense records": True,
    }


def test_validate_workbook_missing_file_fails_every_check():
    result = ingest.validate_workbook(str("no_such_workbook.xlsx"), sheet=config.SHEET_NAME)
    assert result["row_count"] == 0
    assert result["has_existing_classifications"] is False
    assert all(c["ok"] is False for c in result["checks"])


# ---------------------------------------------------------------------------
# Company-norms context (Stream B) — prompts.load_company_norms + its optional
# system block in build_system_prompt. Phase 1 (build_sweep_system_prompt)
# deliberately does NOT get this block (unresolved operator gate — see
# prompts.py's module docstring).
# ---------------------------------------------------------------------------

def test_load_company_norms_missing_file_is_none(monkeypatch, tmp_path):
    from gna_pipeline import prompts

    monkeypatch.setattr(config, "COMPANY_NORMS_MD", tmp_path / "does_not_exist.md")
    assert prompts.load_company_norms() is None


def test_load_company_norms_empty_file_is_none(monkeypatch, tmp_path):
    from gna_pipeline import prompts

    path = tmp_path / "companynorm.md"
    path.write_text("   \n\n", encoding="utf-8")
    monkeypatch.setattr(config, "COMPANY_NORMS_MD", path)
    assert prompts.load_company_norms() is None


def test_load_company_norms_comments_only_is_none(monkeypatch, tmp_path):
    from gna_pipeline import prompts

    path = tmp_path / "companynorm.md"
    path.write_text("<!-- operator instructions, never sent to the model -->\n", encoding="utf-8")
    monkeypatch.setattr(config, "COMPANY_NORMS_MD", path)
    assert prompts.load_company_norms() is None


def test_load_company_norms_returns_stripped_text(monkeypatch, tmp_path):
    from gna_pipeline import prompts

    path = tmp_path / "companynorm.md"
    path.write_text(
        "<!-- instructions -->\n  Acme Corp is a routine monthly vendor.  \n",
        encoding="utf-8",
    )
    monkeypatch.setattr(config, "COMPANY_NORMS_MD", path)
    assert prompts.load_company_norms() == "Acme Corp is a routine monthly vendor."


def test_build_system_prompt_includes_norms_block_only_when_non_empty(monkeypatch, tmp_path):
    from gna_pipeline import prompts

    missing = tmp_path / "does_not_exist.md"
    monkeypatch.setattr(config, "COMPANY_NORMS_MD", missing)
    texts_without = [b["text"] for b in prompts.build_system_prompt(None, None)]
    assert not any("COMPANY NORMS" in t for t in texts_without)

    path = tmp_path / "companynorm.md"
    path.write_text("Acme Corp is a routine monthly vendor.", encoding="utf-8")
    monkeypatch.setattr(config, "COMPANY_NORMS_MD", path)
    texts_with = [b["text"] for b in prompts.build_system_prompt(None, None)]
    assert any(
        "COMPANY NORMS" in t and "Acme Corp is a routine monthly vendor." in t
        for t in texts_with
    )


def test_build_sweep_system_prompt_never_includes_norms_block(monkeypatch, tmp_path):
    """Phase 1 (the sweep) does not read company-norms context at all --
    build_sweep_system_prompt has no call to load_company_norms."""
    from gna_pipeline import prompts

    path = tmp_path / "companynorm.md"
    path.write_text("Acme Corp is a routine monthly vendor.", encoding="utf-8")
    monkeypatch.setattr(config, "COMPANY_NORMS_MD", path)

    texts = [b["text"] for b in prompts.build_sweep_system_prompt(None)]
    assert not any("COMPANY NORMS" in t for t in texts)
    assert not any("Acme Corp is a routine monthly vendor." in t for t in texts)
