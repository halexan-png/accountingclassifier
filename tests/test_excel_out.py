"""tests/test_excel_out.py — the "Deal Profile" sheet in the output workbook.

A run that built (or reused) a deal profile must publish it for audit: one row
per deal entry with the FULL evidence quotes (the prompt context the classifier
saw is capped, the audit sheet is not). A run with no profile must not add an
empty sheet.
"""

from __future__ import annotations

from openpyxl import Workbook, load_workbook

from gna_pipeline import config, excel_out, reporting


def _profile(evidence_count: int = 5) -> dict:
    return {
        "period_range": "202501-202503",
        "source_acctnum": config.MA_ACCTNUM,
        "entries": [
            {
                "name": "Project Alpha",
                "aliases": ["Alpha"],
                "type": "acquisition",
                "properties": ["123 Main St"],
                "entityids": ["E1"],
                "advisors_seen": ["Firm LLP"],
                "evidence": [
                    {"ref": f"row {i}", "quote": f"quote {i}"}
                    for i in range(evidence_count)
                ],
                "supporting_rows": evidence_count,
                "strength": "normal",
            }
        ],
    }


def _source_workbook(tmp_path):
    """A minimal source .xlsx with the sheet name the pipeline expects."""
    source = tmp_path / "source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = config.SHEET_NAME
    ws.cell(row=1, column=1, value="acctnum")
    wb.save(source)
    return source


def test_write_workbook_adds_deal_profile_sheet(tmp_path):
    source = _source_workbook(tmp_path)
    out = tmp_path / "classified.xlsx"
    summary = reporting.build_summary([], total_rows_in=0)
    ok = excel_out.write_workbook(source, [], summary, out, deal_profile=_profile(3))
    assert ok

    result = load_workbook(out)
    assert excel_out.DEAL_PROFILE_SHEET in result.sheetnames
    ws = result[excel_out.DEAL_PROFILE_SHEET]
    assert ws.cell(row=2, column=1).value == "name"
    assert ws.cell(row=3, column=1).value == "Project Alpha"

    # Full (uncapped) evidence lands in the sheet; locate the column by header
    # so column reordering doesn't silently break the assertion.
    headers = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
    assert "evidence" in headers
    assert ws.cell(row=3, column=headers["evidence"]).value.count("quote") == 3


def test_write_workbook_without_profile_skips_sheet(tmp_path):
    source = _source_workbook(tmp_path)
    out = tmp_path / "classified.xlsx"
    summary = reporting.build_summary([], total_rows_in=0)
    assert excel_out.write_workbook(source, [], summary, out, deal_profile=None)
    assert excel_out.DEAL_PROFILE_SHEET not in load_workbook(out).sheetnames
