"""flatten_q2.py — pre-process the two Q2 workbooks into ONE flat .xlsx
sheet in the exact format gna_pipeline/ingest.py already ingests, so the
proven pipeline runs UNCHANGED downstream of this module.

Q2 ships as two separate workbooks:
  - the G&A workbook: one tab per MRI account (A1 = account name, A2 = the
    MRI account number e.g. "MR70000000"), each tab carrying its own header
    row (position varies — see _find_header_row) and its own column layout
    (e.g. T&E puts Amount in column F, not E).
  - the A&T (Acquisition & Transaction) workbook: already the canonical flat
    format (header row 1), real multi-currency rows (OCURRCODE/Exchange
    Rate/USD Amount are NOT all USD/1/equal like G&A rows). It is NOT a
    clean data block, though: the real file carries thousands of fully-blank
    trailing rows alongside the ~214 real deal rows, so ACCTNUM ==
    config.MA_ACCTNUM is the actual inclusion filter, not "every row".

`flatten()` is the pure core: no IO beyond reading the two input paths, no
network, no cost. It never raises on one bad tab/row — a bad tab is skipped
and counted in `stats["tabs_skipped"]`, a blank-period child row is dropped
and counted, and everything is reported back so the caller (or an operator)
can see exactly what happened. Workbook IO dispatches on file extension
(pyxlsb for .xlsb, openpyxl read_only/data_only for .xlsx) mirroring
ingest.py's dispatch pattern — but iterating EVERY sheet, since ingest.py
picks exactly one. That row-reading pattern is copied here, not imported
from ingest.py: this module owns its own IO and never modifies ingest.py.

Output columns are FROZEN (exact order, exact header text) — see
OUTPUT_HEADERS. The amount header must read exactly "USD Amount" because
ingest._EXPECTED_AMOUNT_HEADER matches it verbatim.
"""

from __future__ import annotations

import argparse
import logging
import re
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook, load_workbook
from pyxlsb import open_workbook

from gna_pipeline import config

logger = logging.getLogger("gna.flatten_q2")

# ---------------------------------------------------------------------------
# Canonical output columns — FROZEN, exact order/text. Column 21 (Source Tab)
# is the only field this module invents; columns 1-20 must not be reordered
# or renamed (ingest.py's amount check is an exact-text match on "USD Amount").
# ---------------------------------------------------------------------------
OUTPUT_HEADERS: list[str] = [
    "PERIOD", "REF", "SOURCE", "ENTITYID", "ACCTNUM", "DEPARTMENT", "AMT",
    "DESCRPN", "PDENTRY", "ENTRDATE", "OCURRCODE", "ADDLDESC", "LASTDATE",
    "USERID", "Category", "Quarter", "Image URL", "Image URL - Hyperlink",
    "Exchange Rate", "USD Amount", "Source Tab",
]

_MRI_RE = re.compile(r"^MR\d+")

# Header-row detection: scan rows 1..12 (1-based) for the first row whose
# lowercased, non-empty cell texts contain ALL three tokens (substring match
# per cell). Falls back to row 6 (the common G&A layout) if none match.
_HEADER_PROBE_TOKENS = ("period", "counterparty", "amount")
_HEADER_PROBE_LIMIT = 12
_HEADER_FALLBACK_ROW = 6  # 1-based

# G&A per-tab column aliases -> canonical field name (case-insensitive
# substring, first alias to match any header wins). Counterparty/Page/Type
# are resolved because the header-row detector expects them on the tab, but
# only the fields actually referenced by the output mapping (below) are read.
_GA_FIELD_ALIASES: dict[str, list[str]] = {
    "quarter": ["quarter"],
    "period": ["period"],
    "date": ["date"],
    "counterparty": ["counterparty"],
    "amount": ["amount"],
    "invoice_link": ["invoice link"],
    "page": ["page"],
    "type": ["type"],
    "userid": ["userid"],
    "descrptn": ["descrpn", "description"],
    "adddesc": ["addldesc", "adddesc"],
}

_AT_SHEET_NAME = "Acquisition & Transaction 2026"


# ---------------------------------------------------------------------------
# Workbook IO — dispatch on extension, normalize both formats to a single
# (sheet_name, dense row-value-lists) iterator. Copied from ingest.py's
# per-sheet reading pattern, but walking every sheet (ingest picks one).
# ---------------------------------------------------------------------------

def _iter_sheets(path: str) -> Iterator[tuple[str, list[list[Any]]]]:
    ext = Path(path).suffix.lower()
    if ext == ".xlsb":
        yield from _iter_sheets_xlsb(path)
    else:
        yield from _iter_sheets_openpyxl(path)


def _iter_sheets_openpyxl(path: str) -> Iterator[tuple[str, list[list[Any]]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for name in wb.sheetnames:
            rows = [list(row) for row in wb[name].iter_rows(values_only=True)]
            yield name, rows
    finally:
        wb.close()


def _iter_sheets_xlsb(path: str) -> Iterator[tuple[str, list[list[Any]]]]:
    wb = open_workbook(path)
    try:
        for name in wb.sheets:
            sh = wb.get_sheet(name)
            try:
                rows: list[list[Any]] = []
                for row in sh.rows():
                    if not row:
                        rows.append([])
                        continue
                    width = max(c.c for c in row) + 1
                    values: list[Any] = [None] * width
                    for c in row:
                        values[c.c] = c.v
                    rows.append(values)
                yield name, rows
            finally:
                sh.close()
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Small cleaners — mirror ingest.py's _clean_id idea (float -> int -> str,
# strip, blank -> None), copied rather than imported so this module owns its
# own IO.
# ---------------------------------------------------------------------------

def _clean_id_like(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def _clean_text_simple(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _first_cell(row: list[Any] | None) -> Any:
    if not row:
        return None
    return row[0]


# ---------------------------------------------------------------------------
# G&A tab processing
# ---------------------------------------------------------------------------

def _find_header_row(rows: list[list[Any]]) -> tuple[int, bool]:
    """Return (0-based row index, used_fallback)."""
    limit = min(_HEADER_PROBE_LIMIT, len(rows))
    for i in range(limit):
        texts = [
            str(v).strip().lower() for v in rows[i] if v is not None and str(v).strip()
        ]
        if all(any(token in t for t in texts) for token in _HEADER_PROBE_TOKENS):
            return i, False
    return _HEADER_FALLBACK_ROW - 1, True


def _build_field_map(header_row: list[Any], aliases: dict[str, list[str]]) -> dict[str, int]:
    texts: dict[int, str] = {}
    for idx, v in enumerate(header_row):
        if v is None:
            continue
        t = str(v).strip().lower()
        if t:
            texts[idx] = t

    field_map: dict[str, int] = {}
    for field, alias_list in aliases.items():
        for alias in alias_list:
            found = next((idx for idx in sorted(texts) if alias in texts[idx]), None)
            if found is not None:
                field_map[field] = found
                break
    return field_map


def _cell(row: list[Any], field_map: dict[str, int], field: str) -> Any:
    idx = field_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _cell_or_blank(row: list[Any], field_map: dict[str, int], field: str) -> Any:
    value = _cell(row, field_map, field)
    return value if value is not None else ""


def _build_ga_output_row(
    data_row: list[Any], field_map: dict[str, int], period: str, mri: str, acct_name: str
) -> list[Any]:
    amount_val = _cell(data_row, field_map, "amount")
    row: list[Any] = [""] * len(OUTPUT_HEADERS)
    row[0] = period                                                  # PERIOD
    row[4] = mri                                                     # ACCTNUM
    row[6] = amount_val if amount_val is not None else ""            # AMT
    row[7] = _cell_or_blank(data_row, field_map, "descrptn")         # DESCRPN
    row[9] = _cell_or_blank(data_row, field_map, "date")             # ENTRDATE
    row[10] = "USD"                                                  # OCURRCODE
    row[11] = _cell_or_blank(data_row, field_map, "adddesc")         # ADDLDESC
    row[13] = _cell_or_blank(data_row, field_map, "userid")          # USERID
    row[14] = acct_name                                              # Category
    row[15] = _cell_or_blank(data_row, field_map, "quarter")         # Quarter
    row[17] = _cell_or_blank(data_row, field_map, "invoice_link")    # Image URL - Hyperlink
    row[18] = 1                                                      # Exchange Rate
    row[19] = amount_val if amount_val is not None else ""           # USD Amount
    row[20] = acct_name                                              # Source Tab
    return row


def _process_ga_workbook(path: str, stats: dict[str, Any]) -> list[list[Any]]:
    output_rows: list[list[Any]] = []
    for name, rows in _iter_sheets(path):
        if len(rows) < 2:
            stats["tabs_skipped"].append(
                {"name": name, "reason": "fewer than 2 rows (no A2 to read)"}
            )
            continue

        raw_a2 = _first_cell(rows[1])
        mri = _clean_id_like(raw_a2)
        if mri is None or not _MRI_RE.match(mri):
            stats["tabs_skipped"].append(
                {"name": name, "reason": f"A2 {raw_a2!r} does not match ^MR\\d+"}
            )
            continue
        acct_name = _clean_text_simple(_first_cell(rows[0])) or name

        hdr_idx, used_fallback = _find_header_row(rows)
        if used_fallback:
            stats["header_detect_warnings"].append(
                f"tab {name!r}: no row in 1..{_HEADER_PROBE_LIMIT} matched "
                f"period/counterparty/amount; used fallback row {_HEADER_FALLBACK_ROW}"
            )
        header_row = rows[hdr_idx] if hdr_idx < len(rows) else []
        field_map = _build_field_map(header_row, _GA_FIELD_ALIASES)

        rows_written = 0
        blank_period_dropped = 0
        for data_row in rows[hdr_idx + 1 :]:
            period = _clean_id_like(_cell(data_row, field_map, "period"))
            if not period:
                blank_period_dropped += 1
                continue
            output_rows.append(_build_ga_output_row(data_row, field_map, period, mri, acct_name))
            rows_written += 1

        stats["tabs_included"].append(
            {
                "name": name,
                "mri": mri,
                "rows_written": rows_written,
                "blank_period_dropped": blank_period_dropped,
            }
        )
    return output_rows


# ---------------------------------------------------------------------------
# A&T tab processing — already the canonical flat format; copy the 20
# canonical columns straight through by header name, preserving native
# currency fields as-is. The real A&T sheet is NOT a clean data block: it
# carries thousands of fully-blank trailing rows, and only the M&A account's
# rows (config.MA_ACCTNUM) are the real deal data -- every other ACCTNUM
# present is out of scope for this pipeline, so ACCTNUM is the actual
# inclusion filter (the blank-row skip below is a courtesy, not the gate).
# ---------------------------------------------------------------------------

def _find_at_sheet(
    sheets: list[tuple[str, list[list[Any]]]]
) -> tuple[str, list[list[Any]]] | None:
    for name, rows in sheets:
        if name.strip().casefold() == _AT_SHEET_NAME.casefold():
            return name, rows
    for name, rows in sheets:
        if not rows:
            continue
        header = rows[0]
        if any(
            v is not None and str(v).strip().casefold() == "usd amount" for v in header
        ):
            return name, rows
    return None


def _at_header_map(header_row: list[Any]) -> dict[str, int]:
    text_to_idx: dict[str, int] = {}
    for idx, v in enumerate(header_row):
        if v is None:
            continue
        text = str(v).strip().lower()
        if text and text not in text_to_idx:
            text_to_idx[text] = idx
    return text_to_idx


def _is_ma_acctnum(value: Any) -> bool:
    """True when a (raw) A&T ACCTNUM cell is config.MA_ACCTNUM, tolerating the
    same float/text variance as everywhere else: cleaned "MR58200000" matches
    directly; a cleaned all-digit value (e.g. a cell that came through as
    58200000.0 / "58200000", without the "MR" prefix) matches by prepending
    "MR" before comparing -- never by comparing digits alone, so an unrelated
    account that merely shares the numeric suffix cannot false-match."""
    cleaned = _clean_id_like(value)
    if cleaned is None:
        return False
    if cleaned == config.MA_ACCTNUM:
        return True
    return cleaned.isdigit() and f"MR{cleaned}" == config.MA_ACCTNUM


def _process_at_workbook(path: str, stats: dict[str, Any]) -> list[list[Any]]:
    sheets = list(_iter_sheets(path))
    found = _find_at_sheet(sheets)
    if found is None:
        stats["header_detect_warnings"].append(
            f"A&T workbook {path!r}: no sheet found with a 'USD Amount' header column"
        )
        return []

    _name, rows = found
    if not rows:
        return []
    header_map = _at_header_map(rows[0])
    at_columns = OUTPUT_HEADERS[:-1]  # everything except Source Tab
    acctnum_idx = header_map.get("acctnum")

    output_rows: list[list[Any]] = []
    for data_row in rows[1:]:
        if all(v is None for v in data_row):
            continue
        acctnum_raw = (
            data_row[acctnum_idx] if acctnum_idx is not None and acctnum_idx < len(data_row) else None
        )
        if not _is_ma_acctnum(acctnum_raw):
            continue
        out_row: list[Any] = []
        for header in at_columns:
            idx = header_map.get(header.strip().lower())
            out_row.append(data_row[idx] if idx is not None and idx < len(data_row) else "")
        out_row.append("A&T")
        output_rows.append(out_row)
    return output_rows


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def flatten(gna_path: str, at_path: str) -> tuple[list[str], list[list[Any]], dict[str, Any]]:
    """Read the two Q2 workbooks, return (headers, rows, stats).

    Pure: no writes, no network, $0. Never raises on one bad tab/row -- see
    module docstring.
    """
    stats: dict[str, Any] = {
        "tabs_included": [],
        "tabs_skipped": [],
        "at_rows": 0,
        "total_rows": 0,
        "header_detect_warnings": [],
    }
    rows = _process_ga_workbook(gna_path, stats)
    at_rows = _process_at_workbook(at_path, stats)
    rows.extend(at_rows)
    stats["at_rows"] = len(at_rows)
    stats["total_rows"] = len(rows)
    return list(OUTPUT_HEADERS), rows, stats


def write_flat_workbook(headers: list[str], rows: list[list[Any]], out_path: str) -> None:
    """Thin .xlsx writer: one sheet named config.SHEET_NAME, row 1 = headers,
    rows 2+ = data."""
    wb = Workbook()
    ws = wb.active
    ws.title = config.SHEET_NAME
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(out_path)


def _format_stats(stats: dict[str, Any]) -> str:
    lines = ["flatten_q2 summary:"]
    lines.append(f"  G&A tabs included: {len(stats['tabs_included'])}")
    for t in stats["tabs_included"]:
        lines.append(
            f"    {t['name']} (acct {t['mri']}): {t['rows_written']} rows written, "
            f"{t['blank_period_dropped']} blank-period rows dropped"
        )
    lines.append(f"  G&A tabs skipped: {len(stats['tabs_skipped'])}")
    for t in stats["tabs_skipped"]:
        lines.append(f"    {t['name']}: {t['reason']}")
    if stats["header_detect_warnings"]:
        lines.append("  Header-detect warnings:")
        for w in stats["header_detect_warnings"]:
            lines.append(f"    {w}")
    lines.append(f"  A&T rows: {stats['at_rows']}")
    lines.append(f"  Total rows written: {stats['total_rows']}")
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m gna_pipeline.flatten_q2",
        description=(
            "Flatten the two Q2 workbooks (multi-tab G&A + flat A&T) into one "
            "flat .xlsx the existing pipeline ingests unchanged."
        ),
    )
    parser.add_argument("gna", help="path to the G&A multi-tab workbook (.xlsb or .xlsx)")
    parser.add_argument("at", help="path to the A&T flat workbook (.xlsb or .xlsx)")
    parser.add_argument("-o", "--output", required=True, help="path to write the flat .xlsx")
    args = parser.parse_args(argv)

    headers, rows, stats = flatten(args.gna, args.at)
    write_flat_workbook(headers, rows, args.output)
    print(_format_stats(stats))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
