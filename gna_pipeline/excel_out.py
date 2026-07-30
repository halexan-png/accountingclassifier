"""excel_out.py — annotate a reconstructed copy of the source data with results.

Owns no aggregation (that's reporting.py) and no API. The ORIGINAL source
workbook is never opened for write and never touched. Because the source is a
binary .xlsb (openpyxl cannot open .xlsb at all, read or write), this module
does values-only reconstruction instead of copy-and-append: it reads the
source's headers + raw row values via ingest.read_raw_rows and writes them
into a brand-new openpyxl Workbook, one sheet, at the same row/column
positions as the source (row i+2 in the source lands on row i+2 here too, so
every record's row_idx still points at the right row). Source formatting and
formulas are NOT carried over — only the raw values plus the appended
pipeline columns.

Two load-bearing behaviors:
  - Lock-safe, never-raises contract: this runs LAST in the pipeline, after
    results.jsonl + summary.json are already durable, so a locked workbook can
    never lose API work. Both file touchpoints are guarded and return False
    (never raise) on a PermissionError/OSError: reading the SOURCE workbook
    (locked by Excel or OneDrive sync — the .xlsb lives in a synced folder)
    and saving the OUT_PATH copy (open in Excel). Each prints recovery
    instructions pointing to `python -m gna_pipeline recover`.
  - Idempotent column append: if the header row already contains the
    appended columns (re-annotating a previously annotated file), overwrite
    those columns in place instead of appending duplicates.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from gna_pipeline import config, ingest
from gna_pipeline.contract import DecisionRecord

# The appended columns, in order. invoice_error carries the fetch/resolution
# failure reason for had_invoice=yes rows whose document went unread — so the
# sheet answers "why wasn't this invoice read" by itself; deal_sweep_status is
# the Phase-1 sweep outcome for M&A rows; missing_info is the human_review
# row's open question.
APPENDED_COLUMNS: tuple[str, ...] = (
    "classification",
    "basis",
    "phase",
    "had_invoice",
    "invoice_accessed",
    "invoice_read",
    "invoice_date",
    "reasoning",
    "evidence",
    "missing_info",
    "invoice_pointer",
    "invoice_error",
    "flags",
    "deal_sweep_status",
)

RUN_SUMMARY_SHEET = "Run Summary"
DEAL_PROFILE_SHEET = "Deal Profile"

# One row per deal entry in the "Deal Profile" sheet, in this column order.
_DEAL_PROFILE_COLUMNS: tuple[str, ...] = (
    "name",
    "supporting_rows",
    "matter_numbers",
    "invoice_numbers",
    "quarters",
    "aliases",
    "properties",
    "entityids",
    "advisors_seen",
    "evidence",
)
_DEAL_PROFILE_COLUMN_WIDTHS: tuple[int, ...] = (28, 14, 18, 18, 16, 24, 28, 18, 28, 80)
_EXCEL_CELL_MAX_CHARS = 32_000  # Excel's hard cell limit is 32,767

_FILL_NON_RECURRING = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_HUMAN_REVIEW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_SKIPPED_NEGATIVE = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_FILL_NOT_PROCESSED = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
# Green ("Good") for reclass — a settled Phase-0 bookkeeping label, distinct
# from the red "Bad" non_recurring shading.
_FILL_RECLASS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
# Light orange on had_invoice/invoice_error when a referenced invoice went
# unread — the auditor's "the deciding document is missing" highlight.
_FILL_INVOICE_UNAVAILABLE = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

_CLASSIFICATION_FILLS: dict[str, PatternFill] = {
    "non_recurring": _FILL_NON_RECURRING,
    "human_review": _FILL_HUMAN_REVIEW,
    "skipped_negative": _FILL_SKIPPED_NEGATIVE,
    "reclass": _FILL_RECLASS,
    # "recurring": intentionally unshaded (no entry).
}

# Header/value bold+shaded style for the Run Summary sheet's section headers.
_SECTION_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _find_or_append_columns(ws: Worksheet) -> dict[str, int]:
    """Resolve the 1-based column index for each of APPENDED_COLUMNS.

    Idempotent: if the header row already names a column (case-insensitive
    exact match), reuse it. Otherwise append a new column after the
    current max, in APPENDED_COLUMNS order, so a fresh file gets all columns
    contiguously and a re-annotated file gets none duplicated.
    """
    header_row = 1
    existing: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue
        text = str(value).strip().lower()
        if text in APPENDED_COLUMNS and text not in existing:
            existing[text] = col

    next_col = ws.max_column + 1
    col_map: dict[str, int] = {}
    for name in APPENDED_COLUMNS:
        if name in existing:
            col_map[name] = existing[name]
        else:
            col_map[name] = next_col
            ws.cell(row=header_row, column=next_col, value=name)
            next_col += 1
    return col_map


def _invoice_pointer(record: DecisionRecord) -> str:
    """The packet's invoice_url if present, else the local doc's
    path + page range when a local document was read, else "" ."""
    invoice_url = record["packet"].get("invoice_url")
    if invoice_url:
        return invoice_url
    invoice = record.get("invoice")
    if invoice and invoice.get("kind") not in ("pdf", "text"):
        return ""  # error summaries carry a failure reason, not a document pointer
    if invoice and invoice.get("source") == "local_file" and invoice.get("path_or_url"):
        pages = invoice.get("pages_read")
        if pages:
            return f"{invoice['path_or_url']} p.{pages}"
        return invoice["path_or_url"]
    return ""


def _deal_sweep_status(record: DecisionRecord) -> str:
    """Per-record Phase-1 sweep outcome for the appended sheet column.

    Only meaningful for phase=="deal_profile" rows (M&A account, auto
    non_recurring); every other row gets "" since the sweep never touched it.
    """
    if record["phase"] != "deal_profile":
        return ""
    flags = record["flags"]
    if "deal_sweep_failed" in flags:
        return "could not gather"
    if "deal_sweep_skipped" in flags:
        return "not swept"
    return "invoice read" if record["invoice_accessed"] == "yes" else "invoice not read"


def _annotate_all_rows(
    ws: Worksheet,
    records: list[DecisionRecord],
    unprocessed_row_idxs: Iterable[int] | None = None,
) -> None:
    col_map = _find_or_append_columns(ws)

    for record in records:
        row = record["row_idx"]
        decision = record["decision"]
        classification = decision.get("classification", "")

        values = {
            "classification": classification,
            "basis": decision.get("basis", ""),
            "phase": record["phase"],
            "had_invoice": record["had_invoice"],
            "invoice_accessed": record["invoice_accessed"],
            # invoice_read is a classify-tool field only; Phase-1 sweep records
            # carry its default ("none_attached"), which would misread as "no
            # invoice" on a swept M&A row whose invoice WAS read — blank it there.
            "invoice_read": (
                decision.get("invoice_read", "") if record["phase"] == "classify" else ""
            ),
            "invoice_date": decision.get("invoice_date") or "",
            "reasoning": decision.get("reasoning", ""),
            "evidence": decision.get("evidence", ""),
            "missing_info": decision.get("missing_info") or "",
            "invoice_pointer": _invoice_pointer(record),
            "invoice_error": (record.get("invoice") or {}).get("error") or "",
            "flags": ",".join(record["flags"]),
            "deal_sweep_status": _deal_sweep_status(record),
        }

        for name, value in values.items():
            ws.cell(row=row, column=col_map[name], value=value)

        class_fill = _CLASSIFICATION_FILLS.get(classification)
        if class_fill is not None:
            ws.cell(row=row, column=col_map["classification"]).fill = class_fill

        if "invoice_unavailable" in record["flags"]:
            ws.cell(row=row, column=col_map["had_invoice"]).fill = _FILL_INVOICE_UNAVAILABLE
            ws.cell(row=row, column=col_map["invoice_error"]).fill = _FILL_INVOICE_UNAVAILABLE

    if unprocessed_row_idxs:
        recorded_rows = {record["row_idx"] for record in records}
        for row in unprocessed_row_idxs:
            if row in recorded_rows:
                continue  # defensive — a row with a real record is never overwritten
            cell = ws.cell(row=row, column=col_map["classification"], value="not_processed")
            cell.fill = _FILL_NOT_PROCESSED


def _write_summary_block(ws: Worksheet, row: int, title: str, value: Any) -> int:
    """Render one summary section starting at `row`; returns the next free row.

    Section header (bold + shaded) followed by key/value rows; a nested dict one
    level deep is rendered as indented key/value rows underneath its own key.
    Simple and readable beats clever.
    """
    header_cell = ws.cell(row=row, column=1, value=title)
    header_cell.font = header_cell.font.copy(bold=True)
    header_cell.fill = _SECTION_HEADER_FILL
    row += 1

    if isinstance(value, dict):
        for key, sub_value in value.items():
            if isinstance(sub_value, dict):
                ws.cell(row=row, column=1, value=f"  {key}")
                row += 1
                for sub_key, sub_sub_value in sub_value.items():
                    ws.cell(row=row, column=2, value=f"    {sub_key}")
                    ws.cell(row=row, column=3, value=str(sub_sub_value))
                    row += 1
            else:
                ws.cell(row=row, column=2, value=str(key))
                ws.cell(row=row, column=3, value=str(sub_value))
                row += 1
    else:
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    return row + 1  # blank spacer row between sections


def _write_run_summary_sheet(wb, summary: dict) -> None:
    if RUN_SUMMARY_SHEET in wb.sheetnames:
        del wb[RUN_SUMMARY_SHEET]
    ws = wb.create_sheet(RUN_SUMMARY_SHEET)

    row = 1
    for key, value in summary.items():
        row = _write_summary_block(ws, row, key, value)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40


def _join_list(values: Any) -> str:
    return "; ".join(str(v) for v in (values or []))


def _evidence_text(evidence: Any) -> str:
    """Full evidence quotes, one per line, truncated only at Excel's cell limit."""
    lines = []
    for ev in evidence or []:
        if isinstance(ev, dict):
            lines.append(f"{ev.get('ref', '?')}: {ev.get('quote', '')}")
        else:
            lines.append(str(ev))
    text = "\n".join(lines)
    if len(text) > _EXCEL_CELL_MAX_CHARS:
        text = text[: _EXCEL_CELL_MAX_CHARS - 3] + "..."
    return text


def _write_deal_profile_sheet(wb, deal_profile: dict) -> None:
    """Human-readable mirror of quarter_deal_profile.json — full evidence lives
    here and in the JSON; the Phase-2 system prompt only gets the compact
    identifier index built by prompts.deal_profile_context_index."""
    if DEAL_PROFILE_SHEET in wb.sheetnames:
        del wb[DEAL_PROFILE_SHEET]
    ws = wb.create_sheet(DEAL_PROFILE_SHEET)

    ws.cell(row=1, column=1, value=f"period_range: {deal_profile.get('period_range', '')}")
    ws.cell(row=1, column=2, value=f"source_acctnum: {deal_profile.get('source_acctnum', '')}")
    ws.cell(row=1, column=3, value=f"quarters: {', '.join(deal_profile.get('quarters', []))}")

    header_row = 2
    for col, name in enumerate(_DEAL_PROFILE_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = cell.font.copy(bold=True)
        cell.fill = _SECTION_HEADER_FILL

    for row, entry in enumerate(deal_profile.get("entries", []), start=header_row + 1):
        ws.cell(row=row, column=1, value=entry.get("name", ""))
        ws.cell(row=row, column=2, value=entry.get("supporting_rows", 0))
        ws.cell(row=row, column=3, value=_join_list(entry.get("matter_numbers")))
        ws.cell(row=row, column=4, value=_join_list(entry.get("invoice_numbers")))
        ws.cell(row=row, column=5, value=_join_list(entry.get("quarters")))
        ws.cell(row=row, column=6, value=_join_list(entry.get("aliases")))
        ws.cell(row=row, column=7, value=_join_list(entry.get("properties")))
        ws.cell(row=row, column=8, value=_join_list(entry.get("entityids")))
        ws.cell(row=row, column=9, value=_join_list(entry.get("advisors_seen")))
        ws.cell(row=row, column=10, value=_evidence_text(entry.get("evidence")))

    for i, width in enumerate(_DEAL_PROFILE_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _reconstruct_source_sheet(wb: Workbook, headers: list[str], raw_rows: list[list]) -> Worksheet:
    """Write `headers` + `raw_rows` (from ingest.read_raw_rows) into `wb`'s
    default sheet at the same row/column positions as the source: row i+2 in
    the source lands on row i+2 here, so every record's row_idx still points
    at the right row. Values only — no source formatting/formulas."""
    ws: Worksheet = wb.active
    ws.title = config.SHEET_NAME[:31]  # Excel's 31-char sheet-name limit

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    for row_offset, row_values in enumerate(raw_rows, start=2):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=row_offset, column=col, value=value)

    _linkify_hyperlink_column(ws, headers, len(raw_rows))
    return ws


def _linkify_hyperlink_column(ws: Worksheet, headers: list[str], n_rows: int) -> None:
    """Best-effort: make 'Image URL - Hyperlink' cells real clickable links.
    Silently does nothing if that header isn't present."""
    try:
        col = headers.index("Image URL - Hyperlink") + 1
    except ValueError:
        return
    for row in range(2, n_rows + 2):
        cell = ws.cell(row=row, column=col)
        url = cell.value
        if isinstance(url, str) and url.strip():
            cell.hyperlink = url.strip()


def _add_human_review_report(wb: Workbook) -> None:
    """Bake the auditor-facing Human Review Report tab into the workbook before
    save, so classified.xlsx ships with it (operator directive: a normal run's
    output should already include the report).

    Cosmetic add-on with a hard rule: any failure here must NOT block the core
    workbook save. The API results are already durable in results.jsonl and the
    annotated sheet + Run Summary + Deal Profile are already built on `wb`, so a
    bug in report layout can never cost API work (never-lose-work contract) —
    it just drops the extra tab and logs why. Imported locally so the report
    module is never on excel_out's import-time critical path.
    """
    try:
        from gna_pipeline import exceloutputsheet
        exceloutputsheet.add_report_sheet(wb)
    except Exception as exc:  # noqa: BLE001 — deliberately swallow; see docstring
        print(
            f"Human Review Report tab skipped ({type(exc).__name__}: {exc}); "
            f"classified.xlsx written without it.",
            flush=True,
        )


def write_workbook(
    source_workbook: Path,
    records: list[DecisionRecord],
    summary: dict,
    out_path: Path,
    deal_profile: dict | None = None,
    *,
    unprocessed_row_idxs: Iterable[int] | None = None,
) -> bool:
    """Annotate a values-only reconstruction of `source_workbook`'s data with
    `records` + `summary`, save to `out_path`. Never opens `source_workbook`
    for write. Returns False (never raises) on a PermissionError/OSError from
    either reading the locked source workbook or saving a locked out_path —
    see module docstring.

    `unprocessed_row_idxs` (e.g. from an interrupted run) marks rows that were
    never attempted: the `classification` column gets the literal string
    "not_processed" with a gray fill, and every other appended column is left
    untouched. Rows that already have a record are skipped defensively.
    Default None leaves existing call sites (e.g. `recover`) unchanged.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        headers, raw_rows = ingest.read_raw_rows(str(source_workbook), sheet=config.SHEET_NAME)
        wb = Workbook()
        ws = _reconstruct_source_sheet(wb, headers, raw_rows)
        _annotate_all_rows(ws, records, unprocessed_row_idxs)
        _write_run_summary_sheet(wb, summary)
        if deal_profile is not None:
            _write_deal_profile_sheet(wb, deal_profile)
        # Auditor-facing Human Review Report tab, built from the annotated sheet
        # just written above (never blocks the save — see helper docstring).
        _add_human_review_report(wb)
        # Open on the annotated G&A sheet — the report is an added tab, not the
        # headline (it lands right after this sheet in the tab order).
        wb.active = wb.sheetnames.index(ws.title)
    except (PermissionError, OSError):
        print(
            f"\nExcel annotation skipped: source workbook '{source_workbook}' could not be read\n"
            f"  (likely open in Excel or locked by OneDrive sync).\n"
            f"  results.jsonl and summary.json are already saved — no data lost.\n"
            f"  Close {source_workbook}, then run:\n"
            f"      python -m gna_pipeline recover",
            flush=True,
        )
        return False

    try:
        wb.save(str(out_path))
        return True
    except PermissionError:
        print(
            f"\nExcel write skipped: '{out_path}' is locked (open in Excel?).\n"
            f"  results.jsonl and summary.json are already saved — no data lost.\n"
            f"  Close {out_path} in Excel, then run:\n"
            f"      python -m gna_pipeline recover",
            flush=True,
        )
        return False

