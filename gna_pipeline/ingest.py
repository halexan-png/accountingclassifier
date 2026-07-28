"""ingest.py — read the source workbook into RowPackets.

Owns workbook IO, column resolution, and field normalization. The source is
now the accounting-processed .xlsb extract
(data/input/G&ARecordsWLink.xlsb); IO dispatches on file extension — pyxlsb
for `.xlsb`, openpyxl (read_only, data_only) for everything else — both
normalized to a single plain-value-list-per-row iterator so the
packet-building logic below is written once.

`read_packets` builds the classification-facing RowPacket list (the only row
view any downstream module sees). `read_raw_rows` returns the workbook's raw
headers/rows verbatim (dates normalized to ISO) for excel_out.py's output
reconstruction. `filter_scope` narrows a RowPacket list to a recent period
window and a materiality floor ahead of Phase-2 classification.

Nothing outside gna_pipeline/contract.py may define row shapes: this module
builds RowPacket instances but does not redefine the TypedDict.
"""

from __future__ import annotations

import datetime
import re
from pathlib import Path
from typing import Any, Iterator, TypedDict

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyxlsb import open_workbook

from gna_pipeline import config
from gna_pipeline.contract import RowPacket

ColumnMap = dict[str, int]

# ---------------------------------------------------------------------------
# Column resolution — case-insensitive substring alias match, PRIORITY-ORDERED
# (first alias that matches ANY header wins), falling back to a hardcoded
# 1-based position when no alias matches. Fallback positions are the
# G&ARecordsWLink.xlsb header table (1-based column order).
#
# Only fields the pipeline actually consumes are listed.
# ---------------------------------------------------------------------------
COLUMN_SPEC: dict[str, tuple[list[str], int]] = {
    "period":      (["period"], 1),
    "ref":         (["ref"], 2),
    "source":      (["source"], 3),
    "entityid":    (["entityid", "entity id"], 4),
    "acctnum":     (["acctnum", "account number", "acct"], 5),
    "department":  (["department", "dept"], 6),
    "descrptn":    (["descrpn", "description", "memo"], 8),
    "entrdate":    (["entrdate", "entry date"], 10),
    "currency":    (["ocurrcode", "currcode", "currency"], 11),
    "adddesc":     (["addldesc", "adddesc"], 12),
    "userid":      (["userid", "user id"], 14),
    "category":    (["category"], 15),
    "invoice_url": (["image url - hyperlink", "hyperlink", "scannedcopyurl", "invoice url"], 18),
    "amount":      (["usd amount", "usdamt"], 20),
}


# The amount field must land on the USD-normalized column specifically — a
# silent fallback to another column would misclassify every row without ever
# tripping the "no alias matched" warning, so it gets its own explicit check.
_EXPECTED_AMOUNT_HEADER = "usd amount"


def resolve_columns(header_values: list[Any]) -> tuple[ColumnMap, list[str]]:
    """Map each COLUMN_SPEC field to a 1-based column index.

    `header_values` is a list of plain header cell values (row 1), not
    openpyxl/pyxlsb cell objects — callers extract `.value`/`.v` first so
    this function works identically for both source formats.

    Returns (col_map, warnings). A warning fires when:
      - no alias matched any header (fell back to the hardcoded position);
      - two distinct fields resolved to the SAME column index (wrong-column
        capture — one of them is reading the other's data);
      - the `amount` field resolved to a header other than exactly
        "USD Amount" (silently falling back to another column is a real
        defect even though an alias "matched").
    """
    headers: dict[int, str] = {}
    for idx, value in enumerate(header_values, start=1):
        text = str(value).strip().lower() if value is not None else ""
        if text:
            headers[idx] = text

    col_map: ColumnMap = {}
    warnings: list[str] = []
    for field, (aliases, fallback) in COLUMN_SPEC.items():
        found = None
        for alias in aliases:
            for idx in sorted(headers):
                if alias in headers[idx]:
                    found = idx
                    break
            if found is not None:
                break
        if found is None:
            warnings.append(
                f"{field}: no header alias matched {aliases!r}; using fallback col {fallback}"
            )
            col_map[field] = fallback
        else:
            col_map[field] = found

    # Wrong-column capture: two fields pointing at the same index.
    by_index: dict[int, list[str]] = {}
    for field, idx in col_map.items():
        by_index.setdefault(idx, []).append(field)
    for idx, fields in by_index.items():
        if len(fields) > 1:
            warnings.append(
                f"column index {idx} (header {headers.get(idx, '<blank>')!r}) resolved for "
                f"MULTIPLE fields {fields!r} - wrong-column capture"
            )

    # amount must land on the exact USD-normalized column, not a fallback alias.
    amount_idx = col_map.get("amount")
    amount_header = headers.get(amount_idx, "")
    if amount_header != _EXPECTED_AMOUNT_HEADER:
        warnings.append(
            f"amount resolved to column {amount_idx} with header {amount_header!r}, "
            f"not the expected {_EXPECTED_AMOUNT_HEADER!r} - check for a fallback to "
            f"another column"
        )

    return col_map, warnings


class ReadStats(TypedDict):
    total_data_rows: int
    skipped_blank_rows: int
    blank_amount_rows: int
    column_warnings: list[str]
    currency_tally: dict[str, int]
    header_row: list[Any]


_NULL_SENTINELS = {"", "NULL", "#NAME?"}

# Plausible Excel serial-date range (roughly 1954-2119); both source formats
# can hand back an un-parsed date as this kind of float/int.
_SERIAL_DATE_MIN = 20000
_SERIAL_DATE_MAX = 80000


def _clean_text(value: Any) -> str | None:
    """Strip padding, drop the workbook's NULL/#NAME? sentinels. Returns None
    for anything that normalizes to empty."""
    if value is None:
        return None
    text = str(value).replace("#NAME?", "").strip()
    if text.upper() in _NULL_SENTINELS or text == "":
        return None
    return text


def _clean_id(value: Any) -> str | None:
    """Normalize an id-like cell (int or text) to a stripped string."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if text.upper() in _NULL_SENTINELS or text == "":
        return None
    return text


def _excel_serial_to_iso(value: float | int) -> str:
    return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(value))).isoformat()


def _clean_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if _SERIAL_DATE_MIN <= value <= _SERIAL_DATE_MAX:
            return _excel_serial_to_iso(value)
    text = str(value).strip()
    if text.upper() in _NULL_SENTINELS or text == "":
        return None
    return text


def _parse_amount(value: Any) -> tuple[float, bool]:
    """Parse a raw amount cell. Bracket-safe: accounting exports sometimes
    render negatives as "(5)" or "(1,234.56)".

    Numeric -> float; string -> strip whitespace/'$'/commas, "(x)" ->
    -float(x); unparseable/blank -> (0.0, blank=True), same semantics as
    amount_was_blank always had.
    """
    if value is None:
        return 0.0, True
    if isinstance(value, bool):
        return 0.0, True
    if isinstance(value, (int, float)):
        return float(value), False

    text = str(value).strip()
    if not text:
        return 0.0, True

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()

    text = text.replace("$", "").replace(",", "").strip()
    if not text:
        return 0.0, True

    try:
        amount = float(text)
    except ValueError:
        return 0.0, True

    return (-amount if negative else amount), False


# ---------------------------------------------------------------------------
# Workbook IO — dispatch on extension, normalize both formats to a single
# plain-value-list-per-row iterator (headers first).
# ---------------------------------------------------------------------------

# Sheet selection is BY COLUMNS, not by tab name. resolve_columns emits a
# "no header alias matched ..." warning once per expected field that had to
# fall back to a hardcoded position -- i.e. a column we could NOT find on that
# sheet -- so counting those warnings is exactly "how many of our columns this
# sheet is missing". We pick the sheet missing the fewest, using the caller's
# preferred name (config.SHEET_NAME) only to break ties. This lets any
# single-sheet export (commonly named "Sheet1") work, while a real multi-sheet
# MRI workbook still lands on its data tab.
def _sheet_match_key(header_values: list[Any], name: str, preferred: str | None) -> tuple[int, int]:
    _col_map, warnings = resolve_columns(header_values)
    missing = sum(1 for w in warnings if "no header alias matched" in w)
    is_preferred = bool(preferred) and name.strip().casefold() == preferred.strip().casefold()
    return (missing, 0 if is_preferred else 1)


def _sheet_headers_openpyxl(path: str) -> dict[str, list[Any]]:
    """First (header) row of every sheet, keyed by sheet name."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        headers: dict[str, list[Any]] = {}
        for name in wb.sheetnames:
            headers[name] = []
            for row in wb[name].iter_rows(values_only=True):
                headers[name] = list(row)
                break
        return headers
    finally:
        wb.close()


def _sheet_headers_xlsb(path: str) -> dict[str, list[Any]]:
    """First (header) row of every sheet, keyed by sheet name."""
    wb = open_workbook(path)
    try:
        headers: dict[str, list[Any]] = {}
        for name in wb.sheets:
            sh = wb.get_sheet(name)
            try:
                headers[name] = []
                for row in sh.rows():
                    width = max((c.c for c in row), default=-1) + 1
                    values: list[Any] = [None] * width
                    for c in row:
                        values[c.c] = c.v
                    headers[name] = values
                    break
            finally:
                sh.close()
        return headers
    finally:
        wb.close()


def _resolve_data_sheet(path: str, preferred: str | None) -> str:
    """Choose the sheet whose header row best matches the expected columns.

    Raises ValueError only if the workbook has no sheets at all. A sheet that
    merely lacks some columns is still chosen (whichever is closest); the
    caller's column-warning checks (read_packets / validate_workbook) then
    report it -- so the actual accept/reject decision stays column-based and
    lives in one place, not here."""
    ext = Path(path).suffix.lower()
    headers = _sheet_headers_xlsb(path) if ext == ".xlsb" else _sheet_headers_openpyxl(path)
    if not headers:
        raise ValueError(f"workbook {path!r} contains no sheets")
    return min(headers, key=lambda name: _sheet_match_key(headers[name], name, preferred))


def _iter_rows_openpyxl(path: str, sheet: str | None) -> Iterator[list[Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws: Worksheet = wb[sheet] if sheet else wb.active
        for row in ws.iter_rows():
            yield [c.value for c in row]
    finally:
        wb.close()


def _resolve_xlsb_sheet(wb: Any, sheet: str | None) -> str:
    names = list(wb.sheets)
    if sheet is None:
        return names[0]
    if sheet in names:
        return sheet
    target = sheet.strip().casefold()
    for name in names:
        if name.strip().casefold() == target:
            return name
    raise ValueError(f"sheet {sheet!r} not found; available sheets: {names!r}")


def _iter_rows_xlsb(path: str, sheet: str | None) -> Iterator[list[Any]]:
    wb = open_workbook(path)
    try:
        sheet_name = _resolve_xlsb_sheet(wb, sheet)
        sh = wb.get_sheet(sheet_name)
        try:
            for row in sh.rows():
                if not row:
                    yield []
                    continue
                width = max(c.c for c in row) + 1
                values: list[Any] = [None] * width
                for c in row:
                    values[c.c] = c.v
                yield values
        finally:
            sh.close()
    finally:
        wb.close()


def _iter_raw_rows(path: str, sheet: str | None) -> Iterator[list[Any]]:
    # `sheet` is a PREFERENCE (tie-breaker), not a hard requirement: the data
    # sheet is resolved BY COLUMNS, so a workbook whose tab is called "Sheet1"
    # is read fine as long as it carries the expected columns.
    ext = Path(path).suffix.lower()
    target = _resolve_data_sheet(path, sheet)
    if ext == ".xlsb":
        yield from _iter_rows_xlsb(path, target)
    else:
        yield from _iter_rows_openpyxl(path, target)


def read_packets(
    path: str, sheet: str | None = None
) -> tuple[list[RowPacket], ReadStats]:
    """Read the workbook into RowPackets and ReadStats.

    row_idx stays aligned to the true 1-based Excel row number — fully blank
    rows are skipped but still advance the counter. `sheet` is a PREFERENCE,
    not a requirement: the data sheet is chosen by column match
    (_resolve_data_sheet), so a workbook whose tab isn't named
    config.SHEET_NAME still reads as long as it carries the expected columns.
    """
    rows_iter = _iter_raw_rows(path, sheet)
    header_row = next(rows_iter)
    col_map, warnings = resolve_columns(header_row)

    def cell(row: list[Any], field: str) -> Any:
        idx = col_map[field] - 1
        return row[idx] if idx < len(row) else None

    packets: list[RowPacket] = []
    blank_amount_rows = 0
    skipped_blank_rows = 0
    currency_tally: dict[str, int] = {}
    excel_row = 1  # header was row 1

    for row in rows_iter:
        excel_row += 1
        if all(v is None for v in row):
            skipped_blank_rows += 1
            continue

        amount, amount_was_blank = _parse_amount(cell(row, "amount"))
        if amount_was_blank:
            blank_amount_rows += 1

        currency = (_clean_text(cell(row, "currency")) or "UNKNOWN").upper()
        currency_tally[currency] = currency_tally.get(currency, 0) + 1

        period = _clean_id(cell(row, "period")) or ""

        packet: RowPacket = RowPacket(
            row_idx=excel_row,
            period=period,
            acctnum=_clean_text(cell(row, "acctnum")) or "",
            ref=_clean_id(cell(row, "ref")) or "",
            entityid=_clean_id(cell(row, "entityid")) or "",
            department=_clean_text(cell(row, "department")),
            category=_clean_text(cell(row, "category")),
            amount=amount,
            amount_was_blank=amount_was_blank,
            currency=currency,
            descrptn=_clean_text(cell(row, "descrptn")),
            adddesc=_clean_text(cell(row, "adddesc")),
            source=_clean_text(cell(row, "source")) or "",
            entrdate=_clean_date(cell(row, "entrdate")),
            userid=_clean_text(cell(row, "userid")),
            invoice_url=_clean_text(cell(row, "invoice_url")),
        )
        packets.append(packet)

    stats = ReadStats(
        total_data_rows=len(packets),
        skipped_blank_rows=skipped_blank_rows,
        blank_amount_rows=blank_amount_rows,
        column_warnings=warnings,
        currency_tally=currency_tally,
        header_row=list(header_row),
    )
    return packets, stats


def read_raw_rows(
    path: str, sheet: str | None = None
) -> tuple[list[str], list[list[Any]]]:
    """Read the workbook's raw headers and rows verbatim, for excel_out.py's
    output reconstruction.

    headers = header strings from row 1, stopping at the last non-empty
    header (20 entries for the real file). rows = one list per SOURCE Excel
    row starting at row 2, INCLUDING fully-blank rows (as [None]*len(headers)),
    each trimmed/padded to len(headers). Excel serial-date floats are
    converted to ISO date strings ONLY in columns whose header contains
    "date" (case-insensitive: ENTRDATE, LASTDATE) — every other value is
    passed through as-is.
    """
    rows_iter = _iter_raw_rows(path, sheet)
    header_row = next(rows_iter)

    headers: list[str] = []
    for value in header_row:
        text = str(value).strip() if value is not None else ""
        if not text:
            break
        headers.append(text)
    n = len(headers)
    date_col_idxs = [i for i, h in enumerate(headers) if "date" in h.lower()]

    rows: list[list[Any]] = []
    for row in rows_iter:
        trimmed: list[Any] = list(row[:n])
        if len(trimmed) < n:
            trimmed.extend([None] * (n - len(trimmed)))
        for i in date_col_idxs:
            trimmed[i] = _row_raw_date(trimmed[i])
        rows.append(trimmed)

    return headers, rows


def _row_raw_date(value: Any) -> Any:
    """Date normalization for read_raw_rows: converts a serial float/int or a
    parsed datetime/date to an ISO string; anything else (None, already a
    string, etc.) passes through unchanged."""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if _SERIAL_DATE_MIN <= value <= _SERIAL_DATE_MAX:
            return _excel_serial_to_iso(value)
    return value


# ---------------------------------------------------------------------------
# Classification scope filter (cli.py calls this ahead of Phase 2).
# ---------------------------------------------------------------------------

_YYYYMM_RE = re.compile(r"^\d{6}$")


def filter_scope(
    packets: list[RowPacket], months: str | None, min_usd: float | None
) -> tuple[list[RowPacket], dict]:
    """Narrow a RowPacket list to a recent period window and a materiality
    floor ahead of Phase-2 classification.

    months: None -> str(config.SCOPE_MONTHS_DEFAULT); "all" (case-insensitive)
    -> no period filter; a positive integer string like "3" -> keep the
    latest N distinct periods present in `packets`; a single bare 6-digit
    YYYYMM token with a valid month like "202607" -> keep exactly that one
    period (a literal period, NOT a count -- see the disambiguation comment
    in the code below); a comma-separated list of 6-digit YYYYMM tokens like
    "202601,202602" -> keep exactly those periods (ValueError naming any
    token not present in the file). Anything else -> ValueError.

    min_usd: None -> config.SCOPE_MIN_USD_DEFAULT; 0 (or below) disables;
    otherwise exclude rows where abs(amount) < min_usd, EXCEPT rows with
    amount_was_blank=True are always kept (a human must see them, never
    silently dropped).

    Period filter is applied first, then the amount filter.
    """
    periods_all = sorted({p["period"] for p in packets if p.get("period")})
    periods_all_set = set(periods_all)

    months_resolved = str(config.SCOPE_MONTHS_DEFAULT) if months is None else months
    months_stripped = months_resolved.strip()
    is_all = months_stripped.lower() == "all"

    if is_all:
        periods_selected = list(periods_all)
    elif "," in months_stripped:
        tokens = [t.strip() for t in months_stripped.split(",") if t.strip()]
        for token in tokens:
            if not _YYYYMM_RE.match(token):
                raise ValueError(
                    f"invalid period token {token!r} in months={months_resolved!r}; "
                    f"expected 6-digit YYYYMM tokens"
                )
            if token not in periods_all_set:
                raise ValueError(
                    f"period {token!r} not present in file; available periods: {periods_all!r}"
                )
        periods_selected = sorted(set(tokens))
    elif _YYYYMM_RE.match(months_stripped) and 1 <= int(months_stripped[4:6]) <= 12:
        # Disambiguation rule: a BARE (comma-free) 6-digit token whose last
        # two digits are a valid month (01-12) is a literal single YYYYMM
        # period, not a count -- e.g. "202607" means "July 2026" (exactly
        # one period), matching the comma-list branch above's semantics for
        # a length-1 list. This must be checked BEFORE the plain-count
        # branch below, since a 6-digit count would otherwise also satisfy
        # `.isdigit()`. A short count like "6" is untouched (only 1 digit,
        # doesn't match `_YYYYMM_RE`'s exactly-6-digits shape) and a 6-digit
        # string that isn't a real YYYYMM (e.g. "100000", month "00") falls
        # through to the count branch unchanged.
        if months_stripped not in periods_all_set:
            raise ValueError(
                f"period {months_stripped!r} not present in file; "
                f"available periods: {periods_all!r}"
            )
        periods_selected = [months_stripped]
    elif months_stripped.isdigit() and int(months_stripped) > 0:
        n = int(months_stripped)
        periods_selected = periods_all[-n:]
    else:
        raise ValueError(
            f"invalid months={months_resolved!r}; expected None, 'all', a positive "
            f"integer count of latest periods, or a comma-separated list of "
            f"6-digit YYYYMM periods"
        )

    if is_all:
        period_kept = packets
        excluded_by_period = 0
    else:
        period_set = set(periods_selected)
        period_kept = [p for p in packets if p.get("period") in period_set]
        excluded_by_period = len(packets) - len(period_kept)

    min_usd_resolved = (
        config.SCOPE_MIN_USD_DEFAULT if min_usd is None else float(min_usd)
    )

    if min_usd_resolved <= 0:
        kept = period_kept
        excluded_by_amount = 0
    else:
        kept = [
            p
            for p in period_kept
            if p.get("amount_was_blank") or abs(p["amount"]) >= min_usd_resolved
        ]
        excluded_by_amount = len(period_kept) - len(kept)

    scope_stats = {
        "total_in": len(packets),
        "periods_all": periods_all,
        "periods_selected": sorted(periods_selected),
        "months_arg": months_resolved,
        "min_usd": min_usd_resolved,
        "excluded_by_period": excluded_by_period,
        "excluded_by_amount": excluded_by_amount,
        "kept": len(kept),
    }
    return kept, scope_stats


# ---------------------------------------------------------------------------
# Workbook validation — cheap ($0), UI upload endpoint (v2 UI handoff §6.1,
# `POST /api/workbook`). Reuses resolve_columns/read_packets verbatim; no
# pipeline change beyond this callable, never a second implementation of
# their logic.
# ---------------------------------------------------------------------------

def validate_workbook(path: str, sheet: str | None = None) -> dict[str, Any]:
    """Pre-flight workbook check: a three-item checklist (File format /
    Worksheet structure / Expense records), a row count, and whether the
    workbook already carries classification columns (the entry point for
    resume-by-reupload, v2 UI handoff §4.6/§6.2).

    Returns `{"checks": [{"label": str, "ok": bool}, ...], "row_count": int,
    "has_existing_classifications": bool}`. Never raises: an unreadable file
    (missing, corrupt, wrong sheet) comes back with every check failed
    instead of an exception, so the server can always render a response.
    """
    try:
        _packets, stats = read_packets(path, sheet=sheet)
    except Exception:  # noqa: BLE001 — an upload validation must never 500
        return {
            "checks": [
                {"label": "File format", "ok": False},
                {"label": "Worksheet structure", "ok": False},
                {"label": "Expense records", "ok": False},
            ],
            "row_count": 0,
            "has_existing_classifications": False,
        }

    from gna_pipeline import excel_out  # local: excel_out imports this module

    header_texts = {
        str(value).strip().lower() for value in stats["header_row"] if value is not None
    }
    has_existing_classifications = excel_out.APPENDED_COLUMNS[0] in header_texts

    return {
        "checks": [
            {"label": "File format", "ok": True},
            {"label": "Worksheet structure", "ok": not stats["column_warnings"]},
            {"label": "Expense records", "ok": stats["total_data_rows"] > 0},
        ],
        "row_count": stats["total_data_rows"],
        "has_existing_classifications": has_existing_classifications,
    }
