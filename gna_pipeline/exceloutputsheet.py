"""exceloutputsheet.py — build the auditor-facing "Human Review Report" tab.

Purpose: take the classifier's own output workbook (classified.xlsx), pull out
just the rows a human still needs to look at, and lay them out so an accountant
can audit the charges with all the evidence on one pretty sheet. Owns no
classification logic and never re-decides anything — it only reformats rows the
classifier already produced. Reasoning and evidence are copied VERBATIM from the
classifier's columns; nothing here rewrites them.

WHAT IT DOES
  1. Reads every row from the classifier's main sheet (config.SHEET_NAME) of an
     already-classified workbook.
  2. Keeps only rows classified "human_review" or "non_recurring" (plus any row
     carrying a malformed/unrecognized code like "L", so data-quality issues
     surface instead of getting silently dropped).
  3. Excludes the M&A account (config.MA_ACCTNUM, "MR58200000"): every row on
     that account is non-recurring by definition and auto-swept by the Phase-1
     deal sweep, so it does not belong in a human-review queue.
  4. Groups the survivors by Vendor / Category, vendor groups ordered by total
     dollar exposure (largest first), rows within a group by entry date. This
     is an exact-string group-by on the Category label, nothing fuzzy.
  5. Writes a new "Human Review Report" sheet into the workbook:
       - title bar + subtitle
       - a SUMMARY TABLE by classification (Non-Recurring / Human Review /
         Total), each with Count + USD totals
       - a two-tier gray super-header/header
       - one row per transaction (grayscale banding per vendor group, dynamic
         row height, classification shown as colored TEXT only — gold for Human
         Review, reddish-gray for Non-Recurring)
       - a grand total row

HOW IT IS USED
  The report tab ships INSIDE classified.xlsx: excel_out.write_workbook calls
  add_report_sheet(wb) on the in-memory workbook right before it saves, so every
  `run` / `recover` produces a classified.xlsx that already contains the Human
  Review Report tab — no separate file, no second open. add_report_sheet is the
  integration seam; it reads the annotated main sheet already in the workbook and
  adds/replaces the report tab in place, leaving the active sheet unchanged.

  The module is also runnable standalone to (re)build just the tab on an existing
  classified.xlsx without a full run (see build_report / __main__).

DIFFERENCES FROM THE ORIGINAL PROTOTYPE (build_human_review_report.py)
  - Wired to the real file/sheet (classified.xlsx / config.SHEET_NAME) instead
    of a placeholder "new_upload.xlsx" / "Sheet1".
  - M&A rows are excluded by ACCOUNT NUMBER (config.MA_ACCTNUM), per operator
    instruction, not by the classifier's basis/classify tags. (On the current
    data the two rules select the identical 85 rows.)
  - Summary totals + grand total are written as COMPUTED VALUES, not live
    Excel formulas, so the file opens showing real numbers with no separate
    recalc/LibreOffice step. This is a formatting-only report snapshot, so
    static values are the right call — and it lets us drop the two hidden
    helper columns the formula version needed.
  - The report tab is baked into classified.xlsx rather than written to a
    separate file (operator directive: a normal run's output should already
    include the report).
  - The visible layout, palette, fonts, column set, banding, colored-text
    classification, row-height logic and print setup are unchanged from the
    prototype.

USAGE (standalone — the pipeline calls add_report_sheet directly)
    python -m gna_pipeline.exceloutputsheet [SRC_XLSX] [OUT_XLSX]

  Both arguments are optional. SRC defaults to the classifier output
  (classified.xlsx); OUT defaults to SRC, i.e. the tab is added in place.

DEPENDENCIES
  openpyxl (already a pipeline dependency).
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from gna_pipeline import config

# ---------------------------------------------------------------------------
# Default source. The canonical classifier output is config.CLASSIFIED_XLSX
# (data/output/results/classified.xlsx); a working copy also lives under
# data/output/tempdata/. Prefer the canonical one when it exists, else fall
# back to tempdata, so the standalone entry point keeps working right after a
# real run. There is no separate OUT default — standalone writes in place.
# ---------------------------------------------------------------------------
_TEMPDATA_SRC = config.DATA_OUT / "tempdata" / "classified.xlsx"


def _default_src() -> Path:
    return config.CLASSIFIED_XLSX if config.CLASSIFIED_XLSX.exists() else _TEMPDATA_SRC

REPORT_SHEET = "Human Review Report"

# Classifications this report surfaces, and the order they appear in the
# summary table. "L" (and anything else unrecognized) is surfaced in the data
# rows but is NOT a summary line — it is a data-quality flag, not a category.
KEEP_CLASSIFICATIONS = ("human_review", "non_recurring", "L")
SUMMARY_ORDER = ("non_recurring", "human_review")

# ---------------------------------------------------------------------------
# Palette: grayscale header + banding, Times New Roman throughout.
# (Carried over verbatim from the prototype — this is the agreed look.)
# ---------------------------------------------------------------------------
FONT_NAME = "Times New Roman"
CHARCOAL = "2E2E2E"
SLATE = "6B6E72"
BAND_A = "FFFFFF"        # white
BAND_B = "EFEFEF"        # light gray
NOTE_FONT_COLOR = "444444"
MUTED = "999999"
BORDER_COLOR = "D9D9D9"
LINK_COLOR = "333333"    # link kept understated/gray, not blue
WHITE = "FFFFFF"
SUMMARY_HEADER_FILL = "F3F3F3"

# Classification is signalled by TEXT COLOR ONLY — the cell keeps the row's band fill.
CLASS_STYLE = {
    "human_review":  ("Human Review", "B8860B"),                 # dark goldenrod / pale-yellow family
    "non_recurring": ("Non-Recurring", "8B5E5A"),                 # reddish-gray
    "L":             ("Needs Reclass (raw: \"L\")", "B03A2E"),    # red flag — data-quality issue upstream
}

# ---------------------------------------------------------------------------
# Column plan — 12 visible columns (label, width, super-header group).
# ---------------------------------------------------------------------------
COLUMNS = [
    ("Quarter", 8, "Transaction"),
    ("Period", 10, "Transaction"),
    ("Entry Date", 12, "Transaction"),
    ("Vendor / Category", 26, "Vendor"),
    ("Description", 26, "Description"),
    ("Additional Description", 24, "Description"),
    ("Amount (USD)", 14, "Amount"),
    ("Local Amount", 14, "Amount"),
    ("Supporting Doc", 22, "Documentation"),
    ("Classification", 18, "AI Review"),
    ("Reasoning", 36, "AI Review"),
    ("Evidence", 28, "AI Review"),
]
NCOLS = len(COLUMNS)
COL_WIDTH_BY_LABEL = {c[0]: c[1] for c in COLUMNS}

AMOUNT_USD_COL = 7       # column G
CLASSIFICATION_COL = 10  # column J


# ---------------------------------------------------------------------------
# Read
# ---------------------------------------------------------------------------
def _read_records(wb, sheet_name: str = config.SHEET_NAME) -> list[dict]:
    """Read the classifier's main sheet into a list of dicts keyed by header."""
    src_ws = wb[sheet_name]
    headers = [c.value for c in src_ws[1]]
    records = []
    for r in range(2, src_ws.max_row + 1):
        rec = {headers[i]: src_ws.cell(r, i + 1).value
               for i in range(len(headers)) if headers[i]}
        records.append(rec)
    return records


def _select_rows(records: list[dict]) -> tuple[list[dict], int]:
    """Filter to Human Review / Non-Recurring, drop the M&A account, sort by
    vendor dollar exposure. Returns (rows, n_ma_excluded)."""
    review_rows = [rec for rec in records
                   if rec.get("classification") in KEEP_CLASSIFICATIONS]

    # Exclude the M&A account — every row here is non_recurring by definition
    # and auto-swept by the deal profile, so it is not a human-review item.
    # Keyed off config.MA_ACCTNUM so it tracks the account, not a literal here.
    before = len(review_rows)
    review_rows = [rec for rec in review_rows
                   if str(rec.get("ACCTNUM")) != config.MA_ACCTNUM]
    n_ma_excluded = before - len(review_rows)

    # Vendor grouping: rank vendors by total USD exposure (largest first).
    vendor_totals: dict[str, float] = {}
    for rec in review_rows:
        v = rec.get("Category") or "(uncategorized)"
        vendor_totals[v] = vendor_totals.get(v, 0) + (rec.get("USD Amount") or 0)
    vendor_rank = {v: i for i, v in enumerate(sorted(vendor_totals, key=lambda k: -vendor_totals[k]))}
    review_rows.sort(key=lambda x: (vendor_rank.get(x.get("Category") or "(uncategorized)", 999),
                                    x.get("ENTRDATE") or ""))
    return review_rows, n_ma_excluded


# ---------------------------------------------------------------------------
# Per-row helpers (verbatim behavior from the prototype)
# ---------------------------------------------------------------------------
def _classify_pointer(v):
    if v is None:
        return ("none", None)
    if isinstance(v, str) and v.startswith("http"):
        return ("url", v)
    if isinstance(v, (int, float)):
        return ("number", v)
    if isinstance(v, str):
        try:
            return ("number", float(v))
        except ValueError:
            return ("text", v)
    return ("text", str(v))


def _supporting_doc(rec):
    """Return (kind, display_text, link_or_None) for the Supporting Doc cell."""
    kind, val = _classify_pointer(rec.get("invoice_pointer"))
    had = (rec.get("had_invoice") or "").lower()
    if kind == "url":
        return "link", "View Invoice", val
    if kind == "number":
        return "text", f"Page {int(val)} (batch PDF)", None
    if kind == "text":
        if val.strip().lower() == "unpaid":
            return "text", "Unpaid — no invoice ref", None
        return "text", val, None
    # kind == "none"
    if had == "no":
        return "text", "No invoice on file", None
    return "text", "Invoice on file — no reference", None


def _estimate_row_height(rec):
    texts = [
        (rec.get("reasoning") or "—", COL_WIDTH_BY_LABEL["Reasoning"]),
        (rec.get("evidence") or "—", COL_WIDTH_BY_LABEL["Evidence"]),
        (rec.get("DESCRPN") or "", COL_WIDTH_BY_LABEL["Description"]),
        (rec.get("ADDLDESC") or "—", COL_WIDTH_BY_LABEL["Additional Description"]),
    ]
    max_lines = 1
    for text, width in texts:
        chars_per_line = max(int(width * 1.15), 5)
        lines = max(1, -(-len(str(text)) // chars_per_line))
        max_lines = max(max_lines, lines)
    height = max_lines * 13 + 10
    return min(max(height, 30), 160)


def _summary_stats(review_rows: list[dict]) -> dict[str, dict[str, float]]:
    """Count + USD total per summary classification, computed in Python (no
    live formulas). USD total = sum of the USD-equivalent 'USD Amount'."""
    stats = {k: {"count": 0, "usd": 0.0} for k in SUMMARY_ORDER}
    for rec in review_rows:
        k = rec.get("classification")
        if k not in stats:
            continue  # e.g. "L" — surfaced in data rows, not a summary line
        s = stats[k]
        s["count"] += 1
        s["usd"] += rec.get("USD Amount") or 0
    return stats


def _quarter_label(review_rows: list[dict]) -> str:
    quarters = sorted({rec.get("Quarter") for rec in review_rows if rec.get("Quarter")})
    if not quarters:
        return ""
    if len(quarters) == 1:
        return quarters[0]
    return f"{quarters[0]} – {quarters[-1]}"


# ---------------------------------------------------------------------------
# Build the sheet
# ---------------------------------------------------------------------------
def _build_sheet(wb, review_rows: list[dict], n_ma_excluded: int) -> None:
    if REPORT_SHEET in wb.sheetnames:
        del wb[REPORT_SHEET]
    # Place the report right after the main data sheet so an auditor lands on
    # the data → report pairing; index clamps if there are fewer sheets.
    ws = wb.create_sheet(REPORT_SHEET, index=min(1, len(wb.sheetnames)))

    thin = Side(style="thin", color=BORDER_COLOR)
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    super_font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    super_fill = PatternFill("solid", fgColor=CHARCOAL)
    super_align = Alignment(horizontal="center", vertical="center")

    header_font = Font(name=FONT_NAME, size=10.5, bold=True, color=WHITE)
    header_fill = PatternFill("solid", fgColor=SLATE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name=FONT_NAME, size=10.5, color="1A1A1A")
    note_font = Font(name=FONT_NAME, size=10, italic=True, color=NOTE_FONT_COLOR)
    muted_font = Font(name=FONT_NAME, size=10, italic=True, color=MUTED)
    vendor_font = Font(name=FONT_NAME, size=10.5, bold=True, color="1A1A1A")
    link_font = Font(name=FONT_NAME, size=10.5, color=LINK_COLOR, underline="single")

    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    # ---- Row layout (computed up front) ----
    TITLE_ROW = 1
    SUBTITLE_ROW = 2
    SUMMARY_HEADER_ROW = 4
    SUMMARY_FIRST_ROW = 5
    SUMMARY_TOTAL_ROW = SUMMARY_FIRST_ROW + len(SUMMARY_ORDER)

    n_items = len(review_rows)
    SUPER_ROW = SUMMARY_FIRST_ROW + len(SUMMARY_ORDER) + 1 + 2  # +1 total row, +2 spacer rows
    HEADER_ROW = SUPER_ROW + 1
    FIRST_DATA_ROW = HEADER_ROW + 1
    LAST_DATA_ROW = FIRST_DATA_ROW + n_items - 1

    # ---- Title ----
    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=NCOLS)
    qlabel = _quarter_label(review_rows)
    tcell = ws.cell(TITLE_ROW, 1, f"Human Review Report — {qlabel}".rstrip(" —"))
    tcell.font = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    tcell.fill = PatternFill("solid", fgColor=CHARCOAL)
    tcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[TITLE_ROW].height = 30

    ws.merge_cells(start_row=SUBTITLE_ROW, start_column=1, end_row=SUBTITLE_ROW, end_column=NCOLS)
    scell = ws.cell(SUBTITLE_ROW, 1,
        "Source: classifier output, filtered to rows classified Human Review or Non-Recurring. "
        f"Excludes {n_ma_excluded} rows on the M&A account {config.MA_ACCTNUM} (non-recurring by "
        "definition, auto-swept by the deal profile). Grouped by Vendor / Category, ordered by total "
        "dollar impact — largest exposure first. Banded rows mark each vendor group. Reasoning and "
        "evidence are the classifier's own, shown verbatim.")
    scell.font = Font(name=FONT_NAME, size=10, italic=True, color="595959")
    scell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[SUBTITLE_ROW].height = 34

    # ---- Summary table (computed values) ----
    # Label spans cols 1:2 (merged); Count/USD follow in cols 3,4.
    SUMMARY_LABEL_COL = 1
    SUMMARY_COUNT_COL = 3
    SUMMARY_USD_COL = 4
    SUMMARY_WIDTH = 4

    def merge_label(row):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    merge_label(SUMMARY_HEADER_ROW)
    for col, label in [(SUMMARY_LABEL_COL, "Classification"), (SUMMARY_COUNT_COL, "Count"),
                       (SUMMARY_USD_COL, "USD Total")]:
        cell = ws.cell(SUMMARY_HEADER_ROW, col, label)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=CHARCOAL)
        cell.alignment = align_center
    for col in range(1, SUMMARY_WIDTH + 1):
        c = ws.cell(SUMMARY_HEADER_ROW, col)
        c.fill = PatternFill("solid", fgColor=SUMMARY_HEADER_FILL)
        c.border = Border(top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))
    ws.row_dimensions[SUMMARY_HEADER_ROW].height = 18

    stats = _summary_stats(review_rows)
    for i, cls_key in enumerate(SUMMARY_ORDER):
        row = SUMMARY_FIRST_ROW + i
        label, color = CLASS_STYLE[cls_key]
        s = stats[cls_key]
        merge_label(row)

        lbl_cell = ws.cell(row, SUMMARY_LABEL_COL, label)
        lbl_cell.font = Font(name=FONT_NAME, size=10.5, bold=True, color=color)
        lbl_cell.alignment = align_left

        cnt_cell = ws.cell(row, SUMMARY_COUNT_COL, s["count"])
        cnt_cell.font = Font(name=FONT_NAME, size=10.5)
        cnt_cell.alignment = align_center

        usd_cell = ws.cell(row, SUMMARY_USD_COL, s["usd"])
        usd_cell.font = Font(name=FONT_NAME, size=10.5)
        usd_cell.number_format = '$#,##0.00'
        usd_cell.alignment = align_right

        for c in range(1, SUMMARY_WIDTH + 1):
            ws.cell(row, c).border = box_border

    merge_label(SUMMARY_TOTAL_ROW)
    tot_lbl = ws.cell(SUMMARY_TOTAL_ROW, SUMMARY_LABEL_COL, "Total")
    tot_lbl.font = Font(name=FONT_NAME, size=10.5, bold=True, color=CHARCOAL)
    tot_lbl.alignment = align_left

    totals = {
        SUMMARY_COUNT_COL: (sum(stats[k]["count"] for k in SUMMARY_ORDER), None),
        SUMMARY_USD_COL: (sum(stats[k]["usd"] for k in SUMMARY_ORDER), '$#,##0.00'),
    }
    for col, (value, fmt) in totals.items():
        cell = ws.cell(SUMMARY_TOTAL_ROW, col, value)
        cell.font = Font(name=FONT_NAME, size=10.5, bold=True, color=CHARCOAL)
        if fmt:
            cell.number_format = fmt
        cell.alignment = align_right if col != SUMMARY_COUNT_COL else align_center
    for c in range(1, SUMMARY_WIDTH + 1):
        ws.cell(SUMMARY_TOTAL_ROW, c).border = Border(top=Side(style="thin", color="999999"))

    # ---- Super-header (grouped section labels) ----
    c = 1
    while c <= NCOLS:
        group = COLUMNS[c - 1][2]
        span_end = c
        while span_end < NCOLS and COLUMNS[span_end][2] == group:
            span_end += 1
        if span_end > c:
            ws.merge_cells(start_row=SUPER_ROW, start_column=c, end_row=SUPER_ROW, end_column=span_end)
        cell = ws.cell(SUPER_ROW, c, group.upper())
        cell.font = super_font
        cell.fill = super_fill
        cell.alignment = super_align
        for cc in range(c, span_end + 1):
            ws.cell(SUPER_ROW, cc).fill = super_fill
            ws.cell(SUPER_ROW, cc).border = box_border
        c = span_end + 1
    ws.row_dimensions[SUPER_ROW].height = 18

    # ---- Column headers ----
    for idx, (label, width, _group) in enumerate(COLUMNS, start=1):
        cell = ws.cell(HEADER_ROW, idx, label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = box_border
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 28

    # ---- Data rows ----
    current_vendor = None
    band_toggle = False
    r = FIRST_DATA_ROW
    for rec in review_rows:
        vendor = rec.get("Category") or "(uncategorized)"
        if vendor != current_vendor:
            band_toggle = not band_toggle
            current_vendor = vendor
        band_fill = PatternFill("solid", fgColor=BAND_B if band_toggle else BAND_A)

        entry_date = rec.get("ENTRDATE")
        try:
            entry_date_val = datetime.strptime(entry_date, "%Y-%m-%d") if isinstance(entry_date, str) else entry_date
        except ValueError:
            entry_date_val = entry_date

        ccy = rec.get("OCURRCODE") or "USD"
        orig_amt = rec.get("AMT")
        local_amount_text = f"{ccy} {orig_amt:,.2f}" if ccy != "USD" and orig_amt is not None else "—"

        cls_key = rec.get("classification")
        cls_label, cls_color = CLASS_STYLE.get(cls_key, (str(cls_key), "444444"))

        row_values = {
            1: rec.get("Quarter"),
            2: rec.get("PERIOD"),
            3: entry_date_val,
            4: vendor,
            5: rec.get("DESCRPN"),
            6: rec.get("ADDLDESC") or "—",
            7: rec.get("USD Amount"),
            8: local_amount_text,
            9: None,   # supporting doc, set separately
            10: cls_label,
            11: rec.get("reasoning") or "—",   # verbatim from the classifier
            12: rec.get("evidence") or "—",    # verbatim from the classifier
        }

        for col_idx in range(1, NCOLS + 1):
            cell = ws.cell(r, col_idx, row_values[col_idx])
            cell.border = box_border
            cell.fill = band_fill
            if col_idx == 1:
                cell.font = Font(name=FONT_NAME, size=10.5, bold=True)
                cell.alignment = align_center
            elif col_idx == 3:
                cell.font = body_font
                cell.alignment = align_center
                cell.number_format = "mm/dd/yyyy"
            elif col_idx == 2:
                cell.font = body_font
                cell.alignment = align_center
            elif col_idx == 4:
                cell.font = vendor_font
                cell.alignment = align_left
            elif col_idx in (5, 6):
                cell.font = body_font
                cell.alignment = align_left
            elif col_idx == 7:
                cell.font = Font(name=FONT_NAME, size=10.5, bold=True, color="1A1A1A")
                cell.alignment = align_right
                cell.number_format = '$#,##0.00'
            elif col_idx == 8:
                cell.font = muted_font if local_amount_text == "—" else body_font
                cell.alignment = align_center
            elif col_idx == 10:
                # Classification: NO fill change — only the text color signals status.
                cell.font = Font(name=FONT_NAME, size=10.5, bold=True, color=cls_color)
                cell.alignment = align_center
            elif col_idx in (11, 12):
                cell.font = note_font
                cell.alignment = align_left

        # Supporting doc column
        kind, text, link = _supporting_doc(rec)
        doc_cell = ws.cell(r, 9)
        doc_cell.border = box_border
        doc_cell.fill = band_fill
        doc_cell.value = text
        if kind == "link":
            doc_cell.hyperlink = link
            doc_cell.font = link_font
        else:
            doc_cell.font = muted_font
        doc_cell.alignment = align_center

        ws.row_dimensions[r].height = _estimate_row_height(rec)
        r += 1

    last_data_row = r - 1
    assert last_data_row == LAST_DATA_ROW, "planned vs actual data range drifted — check SUMMARY_ORDER/spacer rows"

    # ---- Grand total ----
    grand_total_usd = sum((rec.get("USD Amount") or 0) for rec in review_rows)
    total_row = r + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=AMOUNT_USD_COL - 1)
    tot_label = ws.cell(total_row, 1, f"Total ({n_items} items flagged Human Review / Non-Recurring)")
    tot_label.font = Font(name=FONT_NAME, size=10.5, bold=True)
    tot_label.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    tot_cell = ws.cell(total_row, AMOUNT_USD_COL, grand_total_usd)
    tot_cell.font = Font(name=FONT_NAME, size=10.5, bold=True)
    tot_cell.number_format = '$#,##0.00'
    tot_cell.alignment = align_right

    for c_ in range(1, NCOLS + 1):
        ws.cell(total_row, c_).border = Border(top=Side(style="thin", color="999999"))

    # ---- Sheet-level settings ----
    ws.freeze_panes = f"B{FIRST_DATA_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(NCOLS)}{last_data_row}"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = SLATE

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = f"{SUPER_ROW}:{HEADER_ROW}"
    # Active-sheet choice is left to the caller: baked into classified.xlsx the
    # workbook keeps opening on the main data sheet (excel_out's existing
    # behavior); the report is an added tab, not the headline.


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------
def add_report_sheet(wb, sheet_name: str = config.SHEET_NAME) -> int:
    """Add/replace the "Human Review Report" tab on an ALREADY-ANNOTATED
    workbook, in memory. Reads `sheet_name` (the classifier's main sheet, which
    must already carry the appended classification/reasoning/evidence columns),
    builds the report tab, and returns the number of rows in the report. The
    workbook is modified in place; the caller is responsible for saving it and
    for the active-sheet choice.

    This is the seam excel_out.write_workbook uses so classified.xlsx ships with
    the report baked in.
    """
    records = _read_records(wb, sheet_name)
    review_rows, n_ma_excluded = _select_rows(records)
    _build_sheet(wb, review_rows, n_ma_excluded)
    return len(review_rows)
def build_report(src_path: Path | str | None = None, out_path: Path | str | None = None) -> bool:
    """Standalone: load an already-classified workbook, add the Human Review
    Report tab, and save. `src_path` defaults to the classifier output
    (classified.xlsx); `out_path` defaults to `src_path`, i.e. the tab is added
    IN PLACE. Use a different `out_path` to write a copy instead.

    Returns False (never raises) if the source can't be read or the output can't
    be saved because a file is open/locked (Excel or OneDrive sync), mirroring
    the pipeline's lock-safe excel_out contract.
    """
    src_path = Path(src_path) if src_path is not None else _default_src()
    out_path = Path(out_path) if out_path is not None else src_path

    if not src_path.exists():
        print(f"Source workbook not found: {src_path}", flush=True)
        return False

    try:
        # Load the whole workbook so all existing tabs are preserved;
        # data_only=False is fine — the source sheets are values-only
        # reconstructions with no formulas to strip.
        wb = openpyxl.load_workbook(src_path, data_only=False)
    except (PermissionError, OSError) as exc:
        print(f"Could not read '{src_path}' (open in Excel / locked by OneDrive?): {exc}", flush=True)
        return False

    if config.SHEET_NAME not in wb.sheetnames:
        print(f"Sheet '{config.SHEET_NAME}' not found in {src_path}. Tabs: {wb.sheetnames}", flush=True)
        return False

    n_rows = add_report_sheet(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out_path)
    except (PermissionError, OSError):
        print(f"\nReport write skipped: '{out_path}' is locked (open in Excel?).\n"
              f"  Close it and re-run: python -m gna_pipeline.exceloutputsheet", flush=True)
        return False

    print(f"Saved: {out_path}  (Human Review Report tab: {n_rows} rows)")
    return True


def main(argv: list[str] | None = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    src = argv[0] if len(argv) >= 1 else None
    out = argv[1] if len(argv) >= 2 else None
    return 0 if build_report(src, out) else 1


if __name__ == "__main__":
    raise SystemExit(main())
